import os
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox
import sys
import json
import logging
import pandas as pd
import openpyxl as pxl
from openpyxl.styles import Font, Color, Border, Side, Alignment, PatternFill, colors
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from math import ceil
from math import log
from math import pi
from wsprops import HSDiag
from wsprops import Visc
from math import pi
from math import sqrt


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1420, 900)
        MainWindow.setMinimumSize(QtCore.QSize(1250, 800))
        MainWindow.setMaximumSize(QtCore.QSize(1680, 1050))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(12)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(50)
        MainWindow.setFont(font)
        icon = QtGui.QIcon()
        path = os.path.join("data", "icon.png")
        icon.addPixmap(QtGui.QPixmap(path), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        MainWindow.setWindowIcon(icon)
        MainWindow.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                 "font: 12pt \"Times New Roman\";")
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.tabWidget = QtWidgets.QTabWidget(self.centralwidget)
        self.tabWidget.setGeometry(QtCore.QRect(10, 10, 1420, 900))
        self.tabWidget.setMinimumSize(QtCore.QSize(1200, 800))
        self.tabWidget.setMaximumSize(QtCore.QSize(1680, 1050))
        self.tabWidget.setObjectName("tabWidget")
        self.tab_3 = QtWidgets.QWidget()
        self.tab_3.setObjectName("tab_3")
        self.groupBox = QtWidgets.QGroupBox(self.tab_3)
        self.groupBox.setGeometry(QtCore.QRect(10, 10, 721, 821))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(12)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(50)
        self.groupBox.setFont(font)
        self.groupBox.setAlignment(QtCore.Qt.AlignCenter)
        self.groupBox.setObjectName("groupBox")
        self.lineEdit_5_obj_name = QtWidgets.QLineEdit(self.groupBox)
        self.lineEdit_5_obj_name.setGeometry(QtCore.QRect(260, 50, 451, 30))
        self.lineEdit_5_obj_name.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_5_obj_name.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_5_obj_name.setObjectName("lineEdit_5_obj_name")
        self.label = QtWidgets.QLabel(self.groupBox)
        self.label.setGeometry(QtCore.QRect(20, 50, 231, 31))
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(self.groupBox)
        self.label_2.setGeometry(QtCore.QRect(20, 110, 400, 50))
        self.label_2.setObjectName("label_2")
        self.lineEdit_6_t_n = QtWidgets.QLineEdit(self.groupBox)
        self.lineEdit_6_t_n.setGeometry(QtCore.QRect(480, 120, 110, 40))
        self.lineEdit_6_t_n.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_6_t_n.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_6_t_n.setObjectName("lineEdit_6_t_n")
        self.label_3 = QtWidgets.QLabel(self.groupBox)
        self.label_3.setGeometry(QtCore.QRect(20, 190, 640, 30))
        self.label_3.setObjectName("label_3")
        self.verticalLayoutWidget = QtWidgets.QWidget(self.groupBox)
        self.verticalLayoutWidget.setGeometry(QtCore.QRect(10, 250, 221, 241))
        self.verticalLayoutWidget.setObjectName("verticalLayoutWidget")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.verticalLayoutWidget)
        self.verticalLayout.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout.setObjectName("verticalLayout")
        self.label_4 = QtWidgets.QLabel(self.verticalLayoutWidget)
        self.label_4.setAlignment(QtCore.Qt.AlignCenter)
        self.label_4.setObjectName("label_4")
        self.verticalLayout.addWidget(self.label_4)
        self.label_6 = QtWidgets.QLabel(self.verticalLayoutWidget)
        self.label_6.setAlignment(QtCore.Qt.AlignCenter)
        self.label_6.setObjectName("label_6")
        self.verticalLayout.addWidget(self.label_6)
        self.label_8 = QtWidgets.QLabel(self.verticalLayoutWidget)
        self.label_8.setAlignment(QtCore.Qt.AlignCenter)
        self.label_8.setObjectName("label_8")
        self.verticalLayout.addWidget(self.label_8)
        self.label_19 = QtWidgets.QLabel(self.verticalLayoutWidget)
        self.label_19.setAlignment(QtCore.Qt.AlignCenter)
        self.label_19.setObjectName("label_19")
        self.verticalLayout.addWidget(self.label_19)
        self.label_9 = QtWidgets.QLabel(self.verticalLayoutWidget)
        self.label_9.setAlignment(QtCore.Qt.AlignCenter)
        self.label_9.setObjectName("label_9")
        self.verticalLayout.addWidget(self.label_9)
        self.verticalLayoutWidget_2 = QtWidgets.QWidget(self.groupBox)
        self.verticalLayoutWidget_2.setGeometry(QtCore.QRect(250, 250, 121, 251))
        self.verticalLayoutWidget_2.setObjectName("verticalLayoutWidget_2")
        self.verticalLayout_2 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_2)
        self.verticalLayout_2.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.lineEdit_11_pol_1 = QtWidgets.QLineEdit(self.verticalLayoutWidget_2)
        self.lineEdit_11_pol_1.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_11_pol_1.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_11_pol_1.setObjectName("lineEdit_11_pol_1")
        self.verticalLayout_2.addWidget(self.lineEdit_11_pol_1)
        self.lineEdit_7_pol_3 = QtWidgets.QLineEdit(self.verticalLayoutWidget_2)
        self.lineEdit_7_pol_3.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_7_pol_3.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_7_pol_3.setObjectName("lineEdit_7_pol_3")
        self.verticalLayout_2.addWidget(self.lineEdit_7_pol_3)
        self.lineEdit_9_ns_16 = QtWidgets.QLineEdit(self.verticalLayoutWidget_2)
        self.lineEdit_9_ns_16.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_9_ns_16.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_9_ns_16.setObjectName("lineEdit_9_ns_16")
        self.verticalLayout_2.addWidget(self.lineEdit_9_ns_16)
        self.lineEdit_9_ns_20 = QtWidgets.QLineEdit(self.verticalLayoutWidget_2)
        self.lineEdit_9_ns_20.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_9_ns_20.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_9_ns_20.setObjectName("lineEdit_9_ns_20")
        self.verticalLayout_2.addWidget(self.lineEdit_9_ns_20)
        self.lineEdit_12_wind = QtWidgets.QLineEdit(self.verticalLayoutWidget_2)
        self.lineEdit_12_wind.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_12_wind.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_12_wind.setObjectName("lineEdit_12_wind")
        self.verticalLayout_2.addWidget(self.lineEdit_12_wind)
        self.verticalLayoutWidget_3 = QtWidgets.QWidget(self.groupBox)
        self.verticalLayoutWidget_3.setGeometry(QtCore.QRect(400, 250, 131, 235))
        self.verticalLayoutWidget_3.setObjectName("verticalLayoutWidget_3")
        self.verticalLayout_3 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_3)
        self.verticalLayout_3.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_3.setObjectName("verticalLayout_3")
        self.label_10 = QtWidgets.QLabel(self.verticalLayoutWidget_3)
        self.label_10.setAlignment(QtCore.Qt.AlignCenter)
        self.label_10.setObjectName("label_10")
        self.verticalLayout_3.addWidget(self.label_10)
        self.label_11 = QtWidgets.QLabel(self.verticalLayoutWidget_3)
        self.label_11.setAlignment(QtCore.Qt.AlignCenter)
        self.label_11.setObjectName("label_11")
        self.verticalLayout_3.addWidget(self.label_11)
        self.label_12 = QtWidgets.QLabel(self.verticalLayoutWidget_3)
        self.label_12.setAlignment(QtCore.Qt.AlignCenter)
        self.label_12.setObjectName("label_12")
        self.verticalLayout_3.addWidget(self.label_12)
        self.label_13 = QtWidgets.QLabel(self.verticalLayoutWidget_3)
        self.label_13.setAlignment(QtCore.Qt.AlignCenter)
        self.label_13.setObjectName("label_13")
        self.verticalLayout_3.addWidget(self.label_13)
        self.lineEdit_another_ok = QtWidgets.QLineEdit(self.verticalLayoutWidget_3)
        self.lineEdit_another_ok.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_another_ok.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_another_ok.setObjectName("lineEdit_another_ok")
        self.verticalLayout_3.addWidget(self.lineEdit_another_ok)
        self.verticalLayoutWidget_4 = QtWidgets.QWidget(self.groupBox)
        self.verticalLayoutWidget_4.setGeometry(QtCore.QRect(550, 250, 131, 251))
        self.verticalLayoutWidget_4.setObjectName("verticalLayoutWidget_4")
        self.verticalLayout_4 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_4)
        self.verticalLayout_4.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_4.setObjectName("verticalLayout_4")
        self.lineEdit_13_pol_2 = QtWidgets.QLineEdit(self.verticalLayoutWidget_4)
        self.lineEdit_13_pol_2.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_13_pol_2.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_13_pol_2.setObjectName("lineEdit_13_pol_2")
        self.verticalLayout_4.addWidget(self.lineEdit_13_pol_2)
        self.lineEdit_14_pol_4 = QtWidgets.QLineEdit(self.verticalLayoutWidget_4)
        self.lineEdit_14_pol_4.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_14_pol_4.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_14_pol_4.setObjectName("lineEdit_14_pol_4")
        self.verticalLayout_4.addWidget(self.lineEdit_14_pol_4)
        self.lineEdit_15_kr = QtWidgets.QLineEdit(self.verticalLayoutWidget_4)
        self.lineEdit_15_kr.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_15_kr.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_15_kr.setObjectName("lineEdit_15_kr")
        self.verticalLayout_4.addWidget(self.lineEdit_15_kr)
        self.lineEdit_16_door = QtWidgets.QLineEdit(self.verticalLayoutWidget_4)
        self.lineEdit_16_door.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_16_door.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_16_door.setObjectName("lineEdit_16_door")
        self.verticalLayout_4.addWidget(self.lineEdit_16_door)
        self.lineEdit_gate = QtWidgets.QLineEdit(self.verticalLayoutWidget_4)
        self.lineEdit_gate.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_gate.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_gate.setObjectName("lineEdit_gate")
        self.verticalLayout_4.addWidget(self.lineEdit_gate)
        self.label_5 = QtWidgets.QLabel(self.groupBox)
        self.label_5.setGeometry(QtCore.QRect(20, 560, 510, 30))
        self.label_5.setObjectName("label_5")
        self.label_7 = QtWidgets.QLabel(self.groupBox)
        self.label_7.setGeometry(QtCore.QRect(49, 610, 21, 30))
        self.label_7.setObjectName("label_7")
        self.comboBox_t1 = QtWidgets.QComboBox(self.groupBox)
        self.comboBox_t1.setGeometry(QtCore.QRect(80, 610, 80, 30))
        self.comboBox_t1.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.comboBox_t1.setObjectName("comboBox_t1")
        self.comboBox_t1.addItem("")
        self.comboBox_t1.addItem("")
        self.comboBox_t1.addItem("")
        self.comboBox_t1.addItem("")
        self.comboBox_t1.addItem("")
        self.comboBox_t1.addItem("")
        self.comboBox_t1.addItem("")
        self.label_14 = QtWidgets.QLabel(self.groupBox)
        self.label_14.setGeometry(QtCore.QRect(200, 610, 20, 30))
        self.label_14.setObjectName("label_14")
        self.comboBox_2_t2 = QtWidgets.QComboBox(self.groupBox)
        self.comboBox_2_t2.setGeometry(QtCore.QRect(230, 610, 80, 30))
        self.comboBox_2_t2.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.comboBox_2_t2.setObjectName("comboBox_2_t2")
        self.comboBox_2_t2.addItem("")
        self.comboBox_2_t2.addItem("")
        self.comboBox_2_t2.addItem("")
        self.label_15 = QtWidgets.QLabel(self.groupBox)
        self.label_15.setGeometry(QtCore.QRect(20, 770, 160, 30))
        self.label_15.setObjectName("label_15")
        self.comboBox_3_tube = QtWidgets.QComboBox(self.groupBox)
        self.comboBox_3_tube.setGeometry(QtCore.QRect(200, 770, 171, 30))
        self.comboBox_3_tube.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.comboBox_3_tube.setObjectName("comboBox_3_tube")
        self.comboBox_3_tube.addItem("")
        self.comboBox_3_tube.addItem("")
        self.checkBox_3_n_coef = QtWidgets.QCheckBox(self.groupBox)
        self.checkBox_3_n_coef.setGeometry(QtCore.QRect(40, 510, 561, 31))
        self.checkBox_3_n_coef.setObjectName("checkBox_3_n_coef")
        self.label_40 = QtWidgets.QLabel(self.groupBox)
        self.label_40.setGeometry(QtCore.QRect(20, 660, 521, 30))
        self.label_40.setObjectName("label_40")
        self.lineEdit_p_n = QtWidgets.QLineEdit(self.groupBox)
        self.lineEdit_p_n.setGeometry(QtCore.QRect(69, 710, 91, 31))
        self.lineEdit_p_n.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_p_n.setText("")
        self.lineEdit_p_n.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_p_n.setObjectName("lineEdit_p_n")
        self.comboBox_4_p_n = QtWidgets.QComboBox(self.groupBox)
        self.comboBox_4_p_n.setGeometry(QtCore.QRect(190, 710, 111, 31))
        self.comboBox_4_p_n.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.comboBox_4_p_n.setObjectName("comboBox_4_p_n")
        self.comboBox_4_p_n.addItem("")
        self.comboBox_4_p_n.addItem("")
        self.comboBox_4_p_n.addItem("")
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(9)
        self.groupBox_2 = QtWidgets.QGroupBox(self.tab_3)
        self.groupBox_2.setGeometry(QtCore.QRect(740, 10, 661, 301))
        self.groupBox_2.setAlignment(QtCore.Qt.AlignCenter)
        self.groupBox_2.setObjectName("groupBox_2")
        self.verticalLayoutWidget_5 = QtWidgets.QWidget(self.groupBox_2)
        self.verticalLayoutWidget_5.setGeometry(QtCore.QRect(10, 50, 431, 231))
        self.verticalLayoutWidget_5.setObjectName("verticalLayoutWidget_5")
        self.verticalLayout_5 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_5)
        self.verticalLayout_5.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_5.setObjectName("verticalLayout_5")
        self.label_16 = QtWidgets.QLabel(self.verticalLayoutWidget_5)
        self.label_16.setObjectName("label_16")
        self.verticalLayout_5.addWidget(self.label_16)
        self.label_18 = QtWidgets.QLabel(self.verticalLayoutWidget_5)
        self.label_18.setObjectName("label_18")
        self.verticalLayout_5.addWidget(self.label_18)
        self.label_21 = QtWidgets.QLabel(self.verticalLayoutWidget_5)
        self.label_21.setObjectName("label_21")
        self.verticalLayout_5.addWidget(self.label_21)
        self.label_22 = QtWidgets.QLabel(self.verticalLayoutWidget_5)
        self.label_22.setObjectName("label_22")
        self.verticalLayout_5.addWidget(self.label_22)
        self.label_23 = QtWidgets.QLabel(self.verticalLayoutWidget_5)
        self.label_23.setObjectName("label_23")
        self.verticalLayout_5.addWidget(self.label_23)
        self.verticalLayoutWidget_6 = QtWidgets.QWidget(self.groupBox_2)
        self.verticalLayoutWidget_6.setGeometry(QtCore.QRect(480, 50, 160, 241))
        self.verticalLayoutWidget_6.setObjectName("verticalLayoutWidget_6")
        self.verticalLayout_6 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_6)
        self.verticalLayout_6.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_6.setObjectName("verticalLayout_6")
        self.lineEdit_17_sum_h_loss = QtWidgets.QLineEdit(self.verticalLayoutWidget_6)
        self.lineEdit_17_sum_h_loss.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_17_sum_h_loss.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_17_sum_h_loss.setReadOnly(True)
        self.lineEdit_17_sum_h_loss.setObjectName("lineEdit_17_sum_h_loss")
        self.verticalLayout_6.addWidget(self.lineEdit_17_sum_h_loss)
        self.lineEdit_18_sum_heat = QtWidgets.QLineEdit(self.verticalLayoutWidget_6)
        self.lineEdit_18_sum_heat.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_18_sum_heat.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_18_sum_heat.setReadOnly(True)
        self.lineEdit_18_sum_heat.setObjectName("lineEdit_18_sum_heat")
        self.verticalLayout_6.addWidget(self.lineEdit_18_sum_heat)
        self.lineEdit_20_sum_rate = QtWidgets.QLineEdit(self.verticalLayoutWidget_6)
        self.lineEdit_20_sum_rate.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_20_sum_rate.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_20_sum_rate.setReadOnly(True)
        self.lineEdit_20_sum_rate.setObjectName("lineEdit_20_sum_rate")
        self.verticalLayout_6.addWidget(self.lineEdit_20_sum_rate)
        self.lineEdit_19_name_gck = QtWidgets.QLineEdit(self.verticalLayoutWidget_6)
        self.lineEdit_19_name_gck.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_19_name_gck.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_19_name_gck.setReadOnly(True)
        self.lineEdit_19_name_gck.setObjectName("lineEdit_19_name_gck")
        self.verticalLayout_6.addWidget(self.lineEdit_19_name_gck)
        self.lineEdit_21_sum_p_loss = QtWidgets.QLineEdit(self.verticalLayoutWidget_6)
        self.lineEdit_21_sum_p_loss.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_21_sum_p_loss.setText("")
        self.lineEdit_21_sum_p_loss.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_21_sum_p_loss.setReadOnly(True)
        self.lineEdit_21_sum_p_loss.setObjectName("lineEdit_21_sum_p_loss")
        self.verticalLayout_6.addWidget(self.lineEdit_21_sum_p_loss)
        self.groupBox_3 = QtWidgets.QGroupBox(self.tab_3)
        self.groupBox_3.setGeometry(QtCore.QRect(740, 320, 661, 261))
        self.groupBox_3.setAlignment(QtCore.Qt.AlignCenter)
        self.groupBox_3.setObjectName("groupBox_3")
        self.verticalLayoutWidget_7 = QtWidgets.QWidget(self.groupBox_3)
        self.verticalLayoutWidget_7.setGeometry(QtCore.QRect(10, 50, 442, 181))
        self.verticalLayoutWidget_7.setObjectName("verticalLayoutWidget_7")
        self.verticalLayout_7 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_7)
        self.verticalLayout_7.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_7.setObjectName("verticalLayout_7")
        self.label_24 = QtWidgets.QLabel(self.verticalLayoutWidget_7)
        self.label_24.setObjectName("label_24")
        self.verticalLayout_7.addWidget(self.label_24)
        self.label_25 = QtWidgets.QLabel(self.verticalLayoutWidget_7)
        self.label_25.setObjectName("label_25")
        self.verticalLayout_7.addWidget(self.label_25)
        self.label_26 = QtWidgets.QLabel(self.verticalLayoutWidget_7)
        self.label_26.setObjectName("label_26")
        self.verticalLayout_7.addWidget(self.label_26)
        self.label_27 = QtWidgets.QLabel(self.verticalLayoutWidget_7)
        self.label_27.setObjectName("label_27")
        self.verticalLayout_7.addWidget(self.label_27)
        self.verticalLayoutWidget_8 = QtWidgets.QWidget(self.groupBox_3)
        self.verticalLayoutWidget_8.setGeometry(QtCore.QRect(480, 50, 160, 181))
        self.verticalLayoutWidget_8.setObjectName("verticalLayoutWidget_8")
        self.verticalLayout_8 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_8)
        self.verticalLayout_8.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_8.setObjectName("verticalLayout_8")
        self.lineEdit_22_sum_n_heat = QtWidgets.QLineEdit(self.verticalLayoutWidget_8)
        self.lineEdit_22_sum_n_heat.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_22_sum_n_heat.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_22_sum_n_heat.setReadOnly(True)
        self.lineEdit_22_sum_n_heat.setObjectName("lineEdit_22_sum_n_heat")
        self.verticalLayout_8.addWidget(self.lineEdit_22_sum_n_heat)
        self.lineEdit_23_dl_heat = QtWidgets.QLineEdit(self.verticalLayoutWidget_8)
        self.lineEdit_23_dl_heat.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_23_dl_heat.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_23_dl_heat.setReadOnly(True)
        self.lineEdit_23_dl_heat.setObjectName("lineEdit_23_dl_heat")
        self.verticalLayout_8.addWidget(self.lineEdit_23_dl_heat)
        self.lineEdit_24_n_rate = QtWidgets.QLineEdit(self.verticalLayoutWidget_8)
        self.lineEdit_24_n_rate.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_24_n_rate.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_24_n_rate.setReadOnly(True)
        self.lineEdit_24_n_rate.setObjectName("lineEdit_24_n_rate")
        self.verticalLayout_8.addWidget(self.lineEdit_24_n_rate)
        self.lineEdit_25_dl_rate = QtWidgets.QLineEdit(self.verticalLayoutWidget_8)
        self.lineEdit_25_dl_rate.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_25_dl_rate.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_25_dl_rate.setReadOnly(True)
        self.lineEdit_25_dl_rate.setObjectName("lineEdit_25_dl_rate")
        self.verticalLayout_8.addWidget(self.lineEdit_25_dl_rate)
        self.pushButton_7_help = QtWidgets.QPushButton(self.tab_3)
        self.pushButton_7_help.setGeometry(QtCore.QRect(760, 740, 171, 71))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(9)
        self.pushButton_7_help.setFont(font)
        self.pushButton_7_help.setStyleSheet("font: 75 14pt \"Times New Roman\";\n"
                                             "background-color: rgb(255, 90, 0);")
        self.pushButton_7_help.setObjectName("pushButton_7_help")
        self.pushButton_save_data = QtWidgets.QPushButton(self.tab_3)
        self.pushButton_save_data.setGeometry(QtCore.QRect(760, 600, 191, 61))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(9)
        self.pushButton_save_data.setFont(font)
        self.pushButton_save_data.setStyleSheet("font: 75 14pt \"Times New Roman\";\n"
                                                "background-color: rgb(0, 170, 255);")
        self.pushButton_save_data.setObjectName("pushButton_save_data")
        self.pushButton_load_data = QtWidgets.QPushButton(self.tab_3)
        self.pushButton_load_data.setGeometry(QtCore.QRect(980, 600, 191, 61))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(9)
        self.pushButton_load_data.setFont(font)
        self.pushButton_load_data.setStyleSheet("font: 75 14pt \"Times New Roman\";\n"
                                                "background-color: rgb(0, 170, 255);")
        self.pushButton_load_data.setObjectName("pushButton_load_data")
        self.pushButton_import = QtWidgets.QPushButton(self.tab_3)
        self.pushButton_import.setGeometry(QtCore.QRect(1200, 600, 191, 61))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(9)
        self.pushButton_import.setFont(font)
        self.pushButton_import.setStyleSheet("font: 75 14pt \"Times New Roman\";\n"
                                             "background-color: rgb(0, 210, 0);")
        self.pushButton_import.setObjectName("pushButton_import")
        self.lineEdit_save_as = QtWidgets.QLineEdit(self.tab_3)
        self.lineEdit_save_as.setGeometry(QtCore.QRect(760, 680, 191, 31))
        self.lineEdit_save_as.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_save_as.setText("")
        self.lineEdit_save_as.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_save_as.setObjectName("lineEdit_save_as")
        self.comboBox_load_from = QtWidgets.QComboBox(self.tab_3)
        self.comboBox_load_from.setGeometry(QtCore.QRect(980, 680, 191, 31))
        self.comboBox_load_from.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.comboBox_load_from.setObjectName("comboBox_load_from")
        self.pushButton_del_files = QtWidgets.QPushButton(self.tab_3)
        self.pushButton_del_files.setGeometry(QtCore.QRect(970, 750, 161, 51))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(9)
        self.pushButton_del_files.setFont(font)
        self.pushButton_del_files.setStyleSheet("font: 75 14pt \"Times New Roman\";\n"
                                                "background-color: rgb(0, 170, 255);")
        self.pushButton_del_files.setObjectName("pushButton_del_files")
        self.lineEdit_save_to_excel = QtWidgets.QLineEdit(self.tab_3)
        self.lineEdit_save_to_excel.setGeometry(QtCore.QRect(1200, 680, 191, 31))
        self.lineEdit_save_to_excel.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_save_to_excel.setText("")
        self.lineEdit_save_to_excel.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_save_to_excel.setObjectName("lineEdit_save_to_excel")
        self.label_20 = QtWidgets.QLabel(self.tab_3)
        self.label_20.setGeometry(QtCore.QRect(1150, 750, 111, 51))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(50)
        self.label_20.setFont(font)
        self.label_20.setStyleSheet("font: 14pt \"Times New Roman\";")
        self.label_20.setAlignment(QtCore.Qt.AlignCenter)
        self.label_20.setObjectName("label_20")
        self.comboBox_wind_style = QtWidgets.QComboBox(self.tab_3)
        self.comboBox_wind_style.setGeometry(QtCore.QRect(1280, 750, 111, 51))
        self.comboBox_wind_style.setStyleSheet("background-color: rgb(85, 255, 255);\n"
                                               "font: 14pt \"Times New Roman\";")
        self.comboBox_wind_style.setObjectName("comboBox_wind_style")
        self.comboBox_wind_style.addItem("")
        self.comboBox_wind_style.addItem("")
        self.comboBox_wind_style.addItem("")
        self.tabWidget.addTab(self.tab_3, "")
        self.tab = QtWidgets.QWidget()
        self.tab.setObjectName("tab")
        self.tableWidget = QtWidgets.QTableWidget(self.tab)
        self.tableWidget.setGeometry(QtCore.QRect(10, 10, 1200, 850))
        self.tableWidget.setMinimumSize(QtCore.QSize(1000, 0))
        self.tableWidget.setMaximumSize(QtCore.QSize(1680, 1050))
        self.tableWidget.setStyleSheet("")
        self.tableWidget.setAlternatingRowColors(True)
        self.tableWidget.setVerticalScrollMode(QtWidgets.QAbstractItemView.ScrollPerItem)
        self.tableWidget.setGridStyle(QtCore.Qt.SolidLine)
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(6)
        self.tableWidget.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(5, item)
        self.tableWidget.horizontalHeader().setDefaultSectionSize(210)
        self.tableWidget.verticalHeader().setCascadingSectionResizes(False)
        self.tableWidget.verticalHeader().setDefaultSectionSize(40)
        self.pushButton_add_row = QtWidgets.QPushButton(self.tab)
        self.pushButton_add_row.setGeometry(QtCore.QRect(1239, 210, 151, 70))
        self.pushButton_add_row.setStyleSheet("background-color: rgb(0, 170, 255);")
        self.pushButton_add_row.setObjectName("pushButton_add_row")
        self.pushButton_2_room = QtWidgets.QPushButton(self.tab)
        self.pushButton_2_room.setGeometry(QtCore.QRect(1240, 40, 151, 70))
        self.pushButton_2_room.setStyleSheet("background-color: rgb(0, 170, 255);")
        self.pushButton_2_room.setObjectName("pushButton_2_room")
        self.pushButton_3_del_row = QtWidgets.QPushButton(self.tab)
        self.pushButton_3_del_row.setGeometry(QtCore.QRect(1240, 400, 151, 70))
        self.pushButton_3_del_row.setStyleSheet("background-color: rgb(0, 170, 255);")
        self.pushButton_3_del_row.setObjectName("pushButton_3_del_row")
        self.lineEdit_f_d_row = QtWidgets.QLineEdit(self.tab)
        self.lineEdit_f_d_row.setGeometry(QtCore.QRect(1230, 500, 70, 40))
        self.lineEdit_f_d_row.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_f_d_row.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_f_d_row.setObjectName("lineEdit_f_d_row")
        self.lineEdit_2_l_d_row = QtWidgets.QLineEdit(self.tab)
        self.lineEdit_2_l_d_row.setGeometry(QtCore.QRect(1330, 500, 70, 40))
        self.lineEdit_2_l_d_row.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_2_l_d_row.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_2_l_d_row.setObjectName("lineEdit_2_l_d_row")
        self.pushButton_4_calc_h_loss = QtWidgets.QPushButton(self.tab)
        self.pushButton_4_calc_h_loss.setGeometry(QtCore.QRect(1240, 760, 151, 70))
        self.pushButton_4_calc_h_loss.setStyleSheet("background-color: rgb(0, 170, 255);")
        self.pushButton_4_calc_h_loss.setObjectName("pushButton_4_calc_h_loss")
        self.lineEdit_3_add_row = QtWidgets.QLineEdit(self.tab)
        self.lineEdit_3_add_row.setGeometry(QtCore.QRect(1280, 300, 70, 40))
        self.lineEdit_3_add_row.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_3_add_row.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_3_add_row.setObjectName("lineEdit_3_add_row")
        self.lineEdit_4_room = QtWidgets.QLineEdit(self.tab)
        self.lineEdit_4_room.setGeometry(QtCore.QRect(1280, 130, 70, 40))
        self.lineEdit_4_room.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_4_room.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_4_room.setObjectName("lineEdit_4_room")
        self.checkBox_inf = QtWidgets.QCheckBox(self.tab)
        self.checkBox_inf.setGeometry(QtCore.QRect(1220, 570, 201, 51))
        self.checkBox_inf.setObjectName("checkBox_inf")
        self.checkBox_2_t_reg = QtWidgets.QCheckBox(self.tab)
        self.checkBox_2_t_reg.setGeometry(QtCore.QRect(1220, 650, 201, 71))
        self.checkBox_2_t_reg.setObjectName("checkBox_2_t_reg")
        self.pushButton_add_row.raise_()
        self.pushButton_2_room.raise_()
        self.tableWidget.raise_()
        self.pushButton_3_del_row.raise_()
        self.lineEdit_f_d_row.raise_()
        self.lineEdit_2_l_d_row.raise_()
        self.pushButton_4_calc_h_loss.raise_()
        self.lineEdit_3_add_row.raise_()
        self.lineEdit_4_room.raise_()
        self.checkBox_inf.raise_()
        self.checkBox_2_t_reg.raise_()
        self.tabWidget.addTab(self.tab, "")
        self.tab_2 = QtWidgets.QWidget()
        self.tab_2.setObjectName("tab_2")
        self.tableWidget_2 = QtWidgets.QTableWidget(self.tab_2)
        self.tableWidget_2.setGeometry(QtCore.QRect(0, 0, 1200, 850))
        self.tableWidget_2.setMinimumSize(QtCore.QSize(1000, 0))
        self.tableWidget_2.setMaximumSize(QtCore.QSize(1680, 1050))
        self.tableWidget_2.setStyleSheet("")
        self.tableWidget_2.setAlternatingRowColors(True)
        self.tableWidget_2.setVerticalScrollMode(QtWidgets.QAbstractItemView.ScrollPerItem)
        self.tableWidget_2.setGridStyle(QtCore.Qt.SolidLine)
        self.tableWidget_2.setObjectName("tableWidget_2")
        self.tableWidget_2.setColumnCount(5)
        self.tableWidget_2.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setHorizontalHeaderItem(4, item)
        self.tableWidget_2.horizontalHeader().setDefaultSectionSize(210)
        self.tableWidget_2.verticalHeader().setCascadingSectionResizes(False)
        self.tableWidget_2.verticalHeader().setDefaultSectionSize(40)
        self.pushButton_form_pr = QtWidgets.QPushButton(self.tab_2)
        self.pushButton_form_pr.setGeometry(QtCore.QRect(1230, 40, 140, 70))
        self.pushButton_form_pr.setStyleSheet("background-color: rgb(0, 170, 255);")
        self.pushButton_form_pr.setObjectName("pushButton_form_pr")
        self.pushButton_del_pr = QtWidgets.QPushButton(self.tab_2)
        self.pushButton_del_pr.setGeometry(QtCore.QRect(1230, 630, 140, 70))
        self.pushButton_del_pr.setStyleSheet("background-color: rgb(0, 170, 255);")
        self.pushButton_del_pr.setObjectName("pushButton_del_pr")
        self.pushButton_calc_pr = QtWidgets.QPushButton(self.tab_2)
        self.pushButton_calc_pr.setGeometry(QtCore.QRect(1230, 750, 140, 70))
        self.pushButton_calc_pr.setStyleSheet("background-color: rgb(0, 170, 255);")
        self.pushButton_calc_pr.setObjectName("pushButton_calc_pr")
        self.comboBox_main_pr_type = QtWidgets.QComboBox(self.tab_2)
        self.comboBox_main_pr_type.setGeometry(QtCore.QRect(1220, 210, 181, 41))
        self.comboBox_main_pr_type.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.comboBox_main_pr_type.setObjectName("comboBox_main_pr_type")
        self.comboBox_main_pr_type.addItem("")
        self.comboBox_main_pr_type.addItem("")
        self.comboBox_main_pr_type.addItem("")
        self.comboBox_main_connect_type = QtWidgets.QComboBox(self.tab_2)
        self.comboBox_main_connect_type.setGeometry(QtCore.QRect(1220, 350, 181, 41))
        self.comboBox_main_connect_type.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.comboBox_main_connect_type.setObjectName("comboBox_main_connect_type")
        self.comboBox_main_connect_type.addItem("")
        self.comboBox_main_connect_type.addItem("")
        self.comboBox_main_connect_type.addItem("")
        self.comboBox_main_connect_type.addItem("")
        self.comboBox_main_connect_type.addItem("")
        self.comboBox_main_connect_type.addItem("")
        self.checkBox_panel = QtWidgets.QCheckBox(self.tab_2)
        self.checkBox_panel.setGeometry(QtCore.QRect(1220, 440, 201, 51))
        self.checkBox_panel.setObjectName("checkBox_panel")
        self.checkBox_k_tr = QtWidgets.QCheckBox(self.tab_2)
        self.checkBox_k_tr.setGeometry(QtCore.QRect(1220, 520, 201, 71))
        self.checkBox_k_tr.setObjectName("checkBox_k_tr")
        self.label_main_pr_type = QtWidgets.QLabel(self.tab_2)
        self.label_main_pr_type.setGeometry(QtCore.QRect(1220, 140, 171, 51))
        self.label_main_pr_type.setStyleSheet("")
        self.label_main_pr_type.setAlignment(QtCore.Qt.AlignCenter)
        self.label_main_pr_type.setObjectName("label_main_pr_type")
        self.label_main_connect_type = QtWidgets.QLabel(self.tab_2)
        self.label_main_connect_type.setGeometry(QtCore.QRect(1220, 280, 171, 51))
        self.label_main_connect_type.setStyleSheet("")
        self.label_main_connect_type.setAlignment(QtCore.Qt.AlignCenter)
        self.label_main_connect_type.setObjectName("label_main_connect_type")
        self.tabWidget.addTab(self.tab_2, "")
        self.tab_4 = QtWidgets.QWidget()
        self.tab_4.setObjectName("tab_4")
        self.tableWidget_3 = QtWidgets.QTableWidget(self.tab_4)
        self.tableWidget_3.setGeometry(QtCore.QRect(0, 0, 1200, 831))
        self.tableWidget_3.setMinimumSize(QtCore.QSize(1000, 0))
        self.tableWidget_3.setMaximumSize(QtCore.QSize(1680, 1050))
        self.tableWidget_3.setStyleSheet("")
        self.tableWidget_3.setAlternatingRowColors(True)
        self.tableWidget_3.setVerticalScrollMode(QtWidgets.QAbstractItemView.ScrollPerItem)
        self.tableWidget_3.setGridStyle(QtCore.Qt.SolidLine)
        self.tableWidget_3.setObjectName("tableWidget_3")
        self.tableWidget_3.setColumnCount(0)
        self.tableWidget_3.setRowCount(18)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_3.setVerticalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_3.setVerticalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_3.setVerticalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_3.setVerticalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_3.setVerticalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_3.setVerticalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_3.setVerticalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_3.setVerticalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_3.setVerticalHeaderItem(8, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_3.setVerticalHeaderItem(9, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_3.setVerticalHeaderItem(10, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_3.setVerticalHeaderItem(11, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_3.setVerticalHeaderItem(12, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_3.setVerticalHeaderItem(13, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_3.setVerticalHeaderItem(14, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_3.setVerticalHeaderItem(15, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_3.setVerticalHeaderItem(16, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_3.setVerticalHeaderItem(17, item)
        self.tableWidget_3.verticalHeader().setDefaultSectionSize(42)
        self.pushButton_add_section = QtWidgets.QPushButton(self.tab_4)
        self.pushButton_add_section.setGeometry(QtCore.QRect(1230, 40, 140, 70))
        self.pushButton_add_section.setStyleSheet("background-color: rgb(0, 170, 255);")
        self.pushButton_add_section.setObjectName("pushButton_add_section")
        self.lineEdit_add_sec = QtWidgets.QLineEdit(self.tab_4)
        self.lineEdit_add_sec.setGeometry(QtCore.QRect(1260, 130, 70, 40))
        self.lineEdit_add_sec.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_add_sec.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_add_sec.setObjectName("lineEdit_add_sec")
        self.pushButton_del_section = QtWidgets.QPushButton(self.tab_4)
        self.pushButton_del_section.setGeometry(QtCore.QRect(1230, 200, 140, 70))
        self.pushButton_del_section.setStyleSheet("background-color: rgb(0, 170, 255);")
        self.pushButton_del_section.setObjectName("pushButton_del_section")
        self.lineEdit_f_d_sec = QtWidgets.QLineEdit(self.tab_4)
        self.lineEdit_f_d_sec.setGeometry(QtCore.QRect(1220, 300, 70, 40))
        self.lineEdit_f_d_sec.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_f_d_sec.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_f_d_sec.setObjectName("lineEdit_f_d_sec")
        self.lineEdit_l_d_sec = QtWidgets.QLineEdit(self.tab_4)
        self.lineEdit_l_d_sec.setGeometry(QtCore.QRect(1320, 300, 70, 40))
        self.lineEdit_l_d_sec.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_l_d_sec.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_l_d_sec.setObjectName("lineEdit_l_d_sec")
        self.pushButton_hydro_calc = QtWidgets.QPushButton(self.tab_4)
        self.pushButton_hydro_calc.setGeometry(QtCore.QRect(1230, 380, 140, 70))
        self.pushButton_hydro_calc.setStyleSheet("background-color: rgb(0, 170, 255);")
        self.pushButton_hydro_calc.setObjectName("pushButton_hydro_calc")
        self.tabWidget.addTab(self.tab_4, "")
        self.tab_5 = QtWidgets.QWidget()
        self.tab_5.setObjectName("tab_5")
        self.tableWidget_4 = QtWidgets.QTableWidget(self.tab_5)
        self.tableWidget_4.setGeometry(QtCore.QRect(10, 10, 1200, 850))
        self.tableWidget_4.setMinimumSize(QtCore.QSize(1000, 0))
        self.tableWidget_4.setMaximumSize(QtCore.QSize(1680, 1050))
        self.tableWidget_4.setStyleSheet("")
        self.tableWidget_4.setAlternatingRowColors(True)
        self.tableWidget_4.setVerticalScrollMode(QtWidgets.QAbstractItemView.ScrollPerItem)
        self.tableWidget_4.setGridStyle(QtCore.Qt.SolidLine)
        self.tableWidget_4.setObjectName("tableWidget_4")
        self.tableWidget_4.setColumnCount(4)
        self.tableWidget_4.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_4.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_4.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_4.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_4.setHorizontalHeaderItem(3, item)
        self.tableWidget_4.horizontalHeader().setDefaultSectionSize(250)
        self.tableWidget_4.verticalHeader().setCascadingSectionResizes(False)
        self.tableWidget_4.verticalHeader().setDefaultSectionSize(40)
        self.pushButton_form_balance = QtWidgets.QPushButton(self.tab_5)
        self.pushButton_form_balance.setGeometry(QtCore.QRect(1240, 40, 140, 70))
        self.pushButton_form_balance.setStyleSheet("background-color: rgb(0, 170, 255);")
        self.pushButton_form_balance.setObjectName("pushButton_form_balance")
        self.pushButton_del_balance = QtWidgets.QPushButton(self.tab_5)
        self.pushButton_del_balance.setGeometry(QtCore.QRect(1240, 160, 140, 70))
        self.pushButton_del_balance.setStyleSheet("background-color: rgb(0, 170, 255);")
        self.pushButton_del_balance.setObjectName("pushButton_del_balance")
        self.pushButton_calc_balance = QtWidgets.QPushButton(self.tab_5)
        self.pushButton_calc_balance.setGeometry(QtCore.QRect(1240, 280, 140, 70))
        self.pushButton_calc_balance.setStyleSheet("background-color: rgb(0, 170, 255);")
        self.pushButton_calc_balance.setObjectName("pushButton_calc_balance")
        self.tabWidget.addTab(self.tab_5, "")
        self.tab_6 = QtWidgets.QWidget()
        self.tab_6.setObjectName("tab_6")
        self.tableWidget_5 = QtWidgets.QTableWidget(self.tab_6)
        self.tableWidget_5.setGeometry(QtCore.QRect(0, 0, 1200, 831))
        self.tableWidget_5.setMinimumSize(QtCore.QSize(1000, 0))
        self.tableWidget_5.setMaximumSize(QtCore.QSize(1680, 1050))
        self.tableWidget_5.setStyleSheet("")
        self.tableWidget_5.setAlternatingRowColors(True)
        self.tableWidget_5.setVerticalScrollMode(QtWidgets.QAbstractItemView.ScrollPerItem)
        self.tableWidget_5.setGridStyle(QtCore.Qt.SolidLine)
        self.tableWidget_5.setObjectName("tableWidget_5")
        self.tableWidget_5.setColumnCount(0)
        self.tableWidget_5.setRowCount(17)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_5.setVerticalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_5.setVerticalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_5.setVerticalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_5.setVerticalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_5.setVerticalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_5.setVerticalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_5.setVerticalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_5.setVerticalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_5.setVerticalHeaderItem(8, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_5.setVerticalHeaderItem(9, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_5.setVerticalHeaderItem(10, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_5.setVerticalHeaderItem(11, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_5.setVerticalHeaderItem(12, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_5.setVerticalHeaderItem(13, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_5.setVerticalHeaderItem(14, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_5.setVerticalHeaderItem(15, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableWidget_5.setVerticalHeaderItem(16, item)
        self.tableWidget_5.verticalHeader().setDefaultSectionSize(42)
        self.pushButton_add_col = QtWidgets.QPushButton(self.tab_6)
        self.pushButton_add_col.setGeometry(QtCore.QRect(1230, 40, 140, 70))
        self.pushButton_add_col.setStyleSheet("background-color: rgb(0, 170, 255);")
        self.pushButton_add_col.setObjectName("pushButton_add_column")
        self.lineEdit_add_col = QtWidgets.QLineEdit(self.tab_6)
        self.lineEdit_add_col.setGeometry(QtCore.QRect(1260, 130, 70, 40))
        self.lineEdit_add_col.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_add_col.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_add_col.setObjectName("lineEdit_add_col")
        self.pushButton_del_col = QtWidgets.QPushButton(self.tab_6)
        self.pushButton_del_col.setGeometry(QtCore.QRect(1230, 200, 140, 70))
        self.pushButton_del_col.setStyleSheet("background-color: rgb(0, 170, 255);")
        self.pushButton_del_col.setObjectName("pushButton_del_col")
        self.label_q_r_choose = QtWidgets.QLabel(self.tab_6)
        self.label_q_r_choose.setGeometry(QtCore.QRect(1220, 360, 171, 51))
        self.label_q_r_choose.setStyleSheet("")
        self.label_q_r_choose.setAlignment(QtCore.Qt.AlignCenter)
        self.label_q_r_choose.setObjectName("label_q_r_choose")
        self.comboBox_q_r_choose = QtWidgets.QComboBox(self.tab_6)
        self.comboBox_q_r_choose.setGeometry(QtCore.QRect(1220, 420, 181, 41))
        self.comboBox_q_r_choose.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.comboBox_q_r_choose.setObjectName("comboBox_q_r_choose")
        self.comboBox_q_r_choose.addItem("")
        self.comboBox_q_r_choose.addItem("")
        self.lineEdit_f_d_col = QtWidgets.QLineEdit(self.tab_6)
        self.lineEdit_f_d_col.setGeometry(QtCore.QRect(1220, 300, 70, 40))
        self.lineEdit_f_d_col.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_f_d_col.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_f_d_col.setObjectName("lineEdit_f_d_col")
        self.lineEdit_l_d_col = QtWidgets.QLineEdit(self.tab_6)
        self.lineEdit_l_d_col.setGeometry(QtCore.QRect(1320, 300, 70, 40))
        self.lineEdit_l_d_col.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_l_d_col.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_l_d_col.setObjectName("lineEdit_l_d_col")
        self.pushButton_tube_calc = QtWidgets.QPushButton(self.tab_6)
        self.pushButton_tube_calc.setGeometry(QtCore.QRect(1230, 500, 140, 70))
        self.pushButton_tube_calc.setStyleSheet("background-color: rgb(0, 170, 255);")
        self.pushButton_tube_calc.setObjectName("pushButton_tube_calc")
        self.label_s_n_sum = QtWidgets.QLabel(self.tab_6)
        self.label_s_n_sum.setGeometry(QtCore.QRect(1210, 600, 80, 40))
        self.label_s_n_sum.setStyleSheet("")
        self.label_s_n_sum.setAlignment(QtCore.Qt.AlignCenter)
        self.label_s_n_sum.setObjectName("label_s_n_sum")
        self.lineEdit_s_n_sum = QtWidgets.QLineEdit(self.tab_6)
        self.lineEdit_s_n_sum.setGeometry(QtCore.QRect(1300, 600, 90, 40))
        self.lineEdit_s_n_sum.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_s_n_sum.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_s_n_sum.setObjectName("lineEdit_s_n_sum")
        self.lineEdit_s_n_sum.setReadOnly(True)
        self.label_s_iz_sum = QtWidgets.QLabel(self.tab_6)
        self.label_s_iz_sum.setGeometry(QtCore.QRect(1210, 660, 80, 40))
        self.label_s_iz_sum.setStyleSheet("")
        self.label_s_iz_sum.setAlignment(QtCore.Qt.AlignCenter)
        self.label_s_iz_sum.setObjectName("label_s_iz_sum")
        self.lineEdit_s_iz_sum = QtWidgets.QLineEdit(self.tab_6)
        self.lineEdit_s_iz_sum.setGeometry(QtCore.QRect(1300, 660, 90, 40))
        self.lineEdit_s_iz_sum.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_s_iz_sum.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_s_iz_sum.setObjectName("lineEdit_s_iz_sum")
        self.lineEdit_s_iz_sum.setReadOnly(True)
        self.label_v_sum = QtWidgets.QLabel(self.tab_6)
        self.label_v_sum.setGeometry(QtCore.QRect(1210, 720, 80, 40))
        self.label_v_sum.setStyleSheet("")
        self.label_v_sum.setAlignment(QtCore.Qt.AlignCenter)
        self.label_v_sum.setObjectName("label_v_sum")
        self.lineEdit_v_sum = QtWidgets.QLineEdit(self.tab_6)
        self.lineEdit_v_sum.setGeometry(QtCore.QRect(1300, 720, 90, 40))
        self.lineEdit_v_sum.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_v_sum.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_v_sum.setObjectName("lineEdit_v_sum")
        self.lineEdit_v_sum.setReadOnly(True)
        self.label_m_sum = QtWidgets.QLabel(self.tab_6)
        self.label_m_sum.setGeometry(QtCore.QRect(1210, 780, 80, 40))
        self.label_m_sum.setStyleSheet("")
        self.label_m_sum.setAlignment(QtCore.Qt.AlignCenter)
        self.label_m_sum.setObjectName("label_m_sum")
        self.lineEdit_m_sum = QtWidgets.QLineEdit(self.tab_6)
        self.lineEdit_m_sum.setGeometry(QtCore.QRect(1300, 780, 90, 40))
        self.lineEdit_m_sum.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.lineEdit_m_sum.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_m_sum.setObjectName("lineEdit_m_sum")
        self.lineEdit_m_sum.setReadOnly(True)
        self.tabWidget.addTab(self.tab_6, "")

        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        self.tabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "   . V.1.1"))
        self.groupBox.setTitle(_translate("MainWindow", " "))
        self.label.setText(_translate("MainWindow", "1.  "))
        self.label_2.setText(_translate("MainWindow", "2.     \n"
                                                      "      , C "))
        self.label_3.setText(_translate("MainWindow",
                                        "<html><head/><body><p>3.    , /<span style=\" vertical-align:super;\">2 </span><span style=\" font-family:\'Times New Roman\',\'serif\'; font-size:14pt;\"> </span>C </p></body></html>"))
        self.label_4.setText(_translate("MainWindow", "-  I "))
        self.label_6.setText(_translate("MainWindow", "-  III "))
        self.label_8.setText(_translate("MainWindow", "-   (<16C)"))
        self.label_19.setText(_translate("MainWindow", "-   (16C)"))
        self.label_9.setText(_translate("MainWindow", "- "))
        self.label_10.setText(_translate("MainWindow", "-  II "))
        self.label_11.setText(_translate("MainWindow", "-  IV "))
        self.label_12.setText(_translate("MainWindow", "- "))
        self.label_13.setText(_translate("MainWindow", "- "))
        self.lineEdit_another_ok.setText(" ")
        self.label_5.setText(_translate("MainWindow", "4.      , C"))
        self.label_7.setText(_translate("MainWindow",
                                        "<html><head/><body><p><span style=\" font-size:14pt;\">t</span><span style=\" font-size:14pt; vertical-align:sub;\">1</span></p></body></html>"))
        self.comboBox_t1.setItemText(0, _translate("MainWindow", "85"))
        self.comboBox_t1.setItemText(1, _translate("MainWindow", "90"))
        self.comboBox_t1.setItemText(2, _translate("MainWindow", "95"))
        self.comboBox_t1.setItemText(3, _translate("MainWindow", "105"))
        self.comboBox_t1.setItemText(4, _translate("MainWindow", "115"))
        self.comboBox_t1.setItemText(5, _translate("MainWindow", "130"))
        self.comboBox_t1.setItemText(6, _translate("MainWindow", "150"))
        self.label_14.setText(_translate("MainWindow",
                                         "<html><head/><body><p><span style=\" font-size:14pt;\">t</span><span style=\" font-size:14pt; vertical-align:sub;\">2</span></p></body></html>"))
        self.comboBox_2_t2.setItemText(0, _translate("MainWindow", "60"))
        self.comboBox_2_t2.setItemText(1, _translate("MainWindow", "65"))
        self.comboBox_2_t2.setItemText(2, _translate("MainWindow", "70"))
        self.label_15.setText(_translate("MainWindow", "6.  "))
        self.comboBox_3_tube.setItemText(0, _translate("MainWindow", ""))
        self.comboBox_3_tube.setItemText(1, _translate("MainWindow", ""))
        self.checkBox_3_n_coef.setText(_translate("MainWindow", "  "))
        self.label_40.setText(_translate("MainWindow", "5.       /  "))
        self.comboBox_4_p_n.setItemText(0, _translate("MainWindow", ""))
        self.comboBox_4_p_n.setItemText(1, _translate("MainWindow", " . ."))
        self.comboBox_4_p_n.setItemText(2, _translate("MainWindow", ""))
        self.groupBox_2.setTitle(_translate("MainWindow", " "))
        self.label_16.setText(_translate("MainWindow", "1.    , "))
        self.label_18.setText(_translate("MainWindow", "2.    , "))
        self.label_21.setText(_translate("MainWindow", "3.     , /"))
        self.label_22.setText(_translate("MainWindow", "4.    ()"))
        self.label_23.setText(_translate("MainWindow", "5.     , "))
        self.groupBox_3.setTitle(_translate("MainWindow", " "))
        self.label_24.setText(_translate("MainWindow", "1.     , "))
        self.label_25.setText(_translate("MainWindow", "2.   , %"))
        self.label_26.setText(_translate("MainWindow", "3.   , /"))
        self.label_27.setText(_translate("MainWindow", "4.   , %"))
        self.pushButton_7_help.setText(_translate("MainWindow", " \n"
                                                                ""))
        self.pushButton_save_data.setText(_translate("MainWindow", "\n"
                                                                   " "))
        self.pushButton_load_data.setText(_translate("MainWindow", "\n"
                                                                   " "))
        self.pushButton_import.setText(_translate("MainWindow", "  \n"
                                                                "  Excel"))
        self.pushButton_del_files.setText(_translate("MainWindow", " "))
        self.label_20.setText(_translate("MainWindow", " \n"
                                                       ""))
        self.comboBox_wind_style.setItemText(0, _translate("MainWindow", "Dark"))
        self.comboBox_wind_style.setItemText(1, _translate("MainWindow", "Medium"))
        self.comboBox_wind_style.setItemText(2, _translate("MainWindow", "Light"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_3), _translate("MainWindow", " "))
        item = self.tableWidget.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", " . ."))
        item = self.tableWidget.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", " "))
        item = self.tableWidget.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", " , "))
        item = self.tableWidget.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", " , "))
        item = self.tableWidget.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", " "))
        item = self.tableWidget.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", ", "))
        self.pushButton_add_row.setText(_translate("MainWindow", " \n"
                                                                 ""))
        self.pushButton_2_room.setText(_translate("MainWindow", " \n"
                                                                ""))
        self.pushButton_3_del_row.setText(_translate("MainWindow", " \n"
                                                                   " "))
        self.pushButton_4_calc_h_loss.setText(_translate("MainWindow", ""))
        self.checkBox_inf.setText(_translate("MainWindow", " \n"
                                                           " (+20%)"))
        self.checkBox_2_t_reg.setText(_translate("MainWindow", " .\n"
                                                               "  60 . 6.2.13\n"
                                                               " (+15%)"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab), _translate("MainWindow", " "))
        item = self.tableWidget_2.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "1"))
        item = self.tableWidget_2.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "2"))
        item = self.tableWidget_2.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "3"))
        item = self.tableWidget_2.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "4"))
        item = self.tableWidget_2.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "5"))
        self.checkBox_panel.setText(_translate("MainWindow", " \n"
                                                             " (+20%)"))
        self.checkBox_k_tr.setText(_translate("MainWindow", " \n"
                                                            ".  (-Q)"))
        self.pushButton_form_pr.setText(_translate("MainWindow", " \n"
                                                                 ""))
        self.pushButton_del_pr.setText(_translate("MainWindow", " \n"
                                                                ""))
        self.pushButton_calc_pr.setText(_translate("MainWindow", ""))
        self.comboBox_main_pr_type.setItemText(0, _translate("MainWindow", ""))
        self.comboBox_main_pr_type.setItemText(1, _translate("MainWindow", ""))
        self.comboBox_main_pr_type.setItemText(2, _translate("MainWindow", "  . ."))
        self.comboBox_main_connect_type.setItemText(0, _translate("MainWindow", ""))
        self.comboBox_main_connect_type.setItemText(1, _translate("MainWindow", " -"))
        self.comboBox_main_connect_type.setItemText(2, _translate("MainWindow", " -"))
        self.comboBox_main_connect_type.setItemText(3, _translate("MainWindow", " -"))
        self.comboBox_main_connect_type.setItemText(4, _translate("MainWindow", " -"))
        self.comboBox_main_connect_type.setItemText(5, _translate("MainWindow", " -"))
        self.label_main_connect_type.setText(_translate("MainWindow", "  \n"
                                                                      " "))
        self.label_main_pr_type.setText(_translate("MainWindow", "  \n"
                                                                 " "))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_2),
                                  _translate("MainWindow", "  "))
        item = self.tableWidget_3.verticalHeaderItem(0)
        item.setText(_translate("MainWindow", ""))
        item = self.tableWidget_3.verticalHeaderItem(1)
        item.setText(_translate("MainWindow", "L, "))
        item = self.tableWidget_3.verticalHeaderItem(2)
        item.setText(_translate("MainWindow", "D, "))
        item = self.tableWidget_3.verticalHeaderItem(3)
        item.setText(_translate("MainWindow", "G, /"))
        item = self.tableWidget_3.verticalHeaderItem(4)
        item.setText(_translate("MainWindow", ". "))
        item = self.tableWidget_3.verticalHeaderItem(5)
        item.setText(_translate("MainWindow", ", "))
        item = self.tableWidget_3.verticalHeaderItem(6)
        item.setText(_translate("MainWindow", "- (), "))
        item = self.tableWidget_3.verticalHeaderItem(7)
        item.setText(_translate("MainWindow", "- (+), "))
        item = self.tableWidget_3.verticalHeaderItem(8)
        item.setText(_translate("MainWindow", ", "))
        item = self.tableWidget_3.verticalHeaderItem(9)
        item.setText(_translate("MainWindow", ", "))
        item = self.tableWidget_3.verticalHeaderItem(10)
        item.setText(_translate("MainWindow", " "))
        item = self.tableWidget_3.verticalHeaderItem(11)
        item.setText(_translate("MainWindow", "w, /"))
        item = self.tableWidget_3.verticalHeaderItem(12)
        item.setText(_translate("MainWindow", "Re"))
        item = self.tableWidget_3.verticalHeaderItem(13)
        item.setText(_translate("MainWindow", ""))
        item = self.tableWidget_3.verticalHeaderItem(14)
        item.setText(_translate("MainWindow", ""))
        item = self.tableWidget_3.verticalHeaderItem(15)
        item.setText(_translate("MainWindow", "S, /(/)^2"))
        item = self.tableWidget_3.verticalHeaderItem(16)
        item.setText(_translate("MainWindow", "p, "))
        item = self.tableWidget_3.verticalHeaderItem(17)
        item.setText(_translate("MainWindow", "p, "))
        self.pushButton_add_section.setText(_translate("MainWindow", "\n"
                                                                     " "))
        self.pushButton_del_section.setText(_translate("MainWindow", " \n"
                                                                     " "))
        self.pushButton_hydro_calc.setText(_translate("MainWindow", ""))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_4), _translate("MainWindow", " "))
        item = self.tableWidget_4.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", " "))
        item = self.tableWidget_4.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", ""))
        item = self.tableWidget_4.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", " , "))
        item = self.tableWidget_4.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "KVs , 3//^0,5"))
        self.pushButton_form_balance.setText(_translate("MainWindow", " \n"
                                                                      ""))
        self.pushButton_del_balance.setText(_translate("MainWindow", " \n"
                                                                     ""))
        self.pushButton_calc_balance.setText(_translate("MainWindow", ""))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_5),
                                  _translate("MainWindow", " "))
        item = self.tableWidget_5.verticalHeaderItem(0)
        item.setText(_translate("MainWindow", " "))
        item = self.tableWidget_5.verticalHeaderItem(1)
        item.setText(_translate("MainWindow", " "))
        item = self.tableWidget_5.verticalHeaderItem(2)
        item.setText(_translate("MainWindow", "Q, "))
        item = self.tableWidget_5.verticalHeaderItem(3)
        item.setText(_translate("MainWindow", "Q, /"))
        item = self.tableWidget_5.verticalHeaderItem(4)
        item.setText(_translate("MainWindow", "G, /"))
        item = self.tableWidget_5.verticalHeaderItem(5)
        item.setText(_translate("MainWindow", "V, /"))
        item = self.tableWidget_5.verticalHeaderItem(6)
        item.setText(_translate("MainWindow", "V, 3/"))
        item = self.tableWidget_5.verticalHeaderItem(7)
        item.setText(_translate("MainWindow", "w, /"))
        item = self.tableWidget_5.verticalHeaderItem(8)
        item.setText(_translate("MainWindow", "d_, "))
        item = self.tableWidget_5.verticalHeaderItem(9)
        item.setText(_translate("MainWindow", "DN, "))
        item = self.tableWidget_5.verticalHeaderItem(10)
        item.setText(_translate("MainWindow", "l, "))
        item = self.tableWidget_5.verticalHeaderItem(11)
        item.setText(_translate("MainWindow", "s_, "))
        item = self.tableWidget_5.verticalHeaderItem(12)
        item.setText(_translate("MainWindow", "S, 2"))
        item = self.tableWidget_5.verticalHeaderItem(13)
        item.setText(_translate("MainWindow", "S, 2"))
        item = self.tableWidget_5.verticalHeaderItem(14)
        item.setText(_translate("MainWindow", "v, 3"))
        item = self.tableWidget_5.verticalHeaderItem(15)
        item.setText(_translate("MainWindow", "M ( ), "))
        item = self.tableWidget_5.verticalHeaderItem(16)
        item.setText(_translate("MainWindow", "M/l ( ), /"))

        self.pushButton_add_col.setText(_translate("MainWindow", "\n"
                                                                 " "))
        self.pushButton_del_col.setText(_translate("MainWindow", " \n"
                                                                 " "))
        self.label_q_r_choose.setText(_translate("MainWindow", ". . "))
        self.comboBox_q_r_choose.setItemText(0, _translate("MainWindow", "Q, "))
        self.comboBox_q_r_choose.setItemText(1, _translate("MainWindow", "Q, /"))
        self.pushButton_tube_calc.setText(_translate("MainWindow", ""))
        self.label_s_n_sum.setText(_translate("MainWindow", "S, 2"))
        self.label_s_iz_sum.setText(_translate("MainWindow", "S, 2"))
        self.label_v_sum.setText(_translate("MainWindow", "v, 3"))
        self.label_m_sum.setText(_translate("MainWindow", "M, "))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_6), _translate("MainWindow", " "))


class Ui_Form(object):
    def setupUi(self, Form):
        Form.setObjectName("Form")
        Form.resize(1000, 1000)
        Form.setMinimumSize(QtCore.QSize(1000, 400))
        Form.setMaximumSize(QtCore.QSize(1000, 1200))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        Form.setFont(font)
        path = os.path.join("data", "icon.png")
        icon = QtGui.QIcon(path)
        Form.setWindowIcon(icon)
        Form.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.textBrowser = QtWidgets.QTextBrowser(Form)
        self.textBrowser.setGeometry(QtCore.QRect(0, 0, 1000, 1000))
        self.textBrowser.setMinimumSize(QtCore.QSize(1000, 400))
        self.textBrowser.setMaximumSize(QtCore.QSize(1000, 1200))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(12)
        self.textBrowser.setFont(font)
        self.textBrowser.setStyleSheet("")
        self.textBrowser.setObjectName("textBrowser")

        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "SVO_Manual"))
        self.textBrowser.setHtml(_translate("Form",
                                            "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
                                            "<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
                                            "p, li { white-space: pre-wrap; }\n"
                                            "</style></head><body style=\" font-family:\'Times New Roman\'; font-size:12pt; font-weight:400; font-style:normal;\">\n"
                                            "<h2 align=\"center\" style=\" margin-top:16px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-size:14pt; font-weight:600;\">    V1.1<br />  </span></h2>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">   ,    .          Calc_SVO_V1.1_Manual.pdf. </p>\n"
                                            "<p align=\"center\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-size:14pt; font-weight:600;\">1.  </span><span style=\" font-size:14pt;\"> </span></p>\n"
                                            "<p align=\"center\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-weight:600;\">1.1   </span></p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-style:italic;\">t</span>,       </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-style:italic;\">Q</span>,         </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-style:italic;\">Q</span>,         </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-style:italic;\">G</span>, /        </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-style:italic;\">Q_</span>,            </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-style:italic;\">Q/</span>,           </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-style:italic;\">N</span>,          </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-style:italic;\">Q</span>,         (   / - ) </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-style:italic;\">L</span>,       </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-style:italic;\">L</span>,        </p>\n"
                                            "<p align=\"center\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-weight:600;\">1.2.  </span> </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-style:italic;\">L</span>,     </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-style:italic;\">D</span>,      </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-style:italic;\">G</span>, /      </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-style:italic;\">- ()</span>,     </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-style:italic;\">- (+)</span>,     (       ) </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-style:italic;\"></span>,      </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-style:italic;\">w</span>, /      </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-style:italic;\">Re</span>      </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-style:italic;\"></span>       </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-style:italic;\">S</span>, /(/)<span style=\" vertical-align:super;\">2</span>     </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'Calibri\',\'sans-serif\'; font-style:italic;\"></span><span style=\" font-style:italic;\">p</span>,       </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-style:italic;\"></span><span style=\" font-family:\'Calibri\',\'sans-serif\'; font-style:italic;\"></span><span style=\" font-style:italic;\">p</span>,           () </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-style:italic;\">KVs</span>, <span style=\" vertical-align:super;\">3</span>//<span style=\" vertical-align:super;\">0,5 </span>      </p>\n"
                                            "<p align=\"center\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-size:14pt; font-weight:600;\">2. </span><span style=\" font-size:14pt;\"> </span></p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-weight:600;\">*</span> <span style=\" font-family:\'Symbol\';\">-</span>         ,   . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-style:italic;\">** </span><span style=\" font-family:\'Symbol\';\">-</span>       <span style=\" font-style:italic;\">SVO_Log.txt</span>. </p>\n"
                                            "<p align=\"center\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-weight:600;\">2.1.  </span> </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">   3 : </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">1. <span style=\" font-style:italic;\"> </span>.      . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">      . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">1.1  <span style=\" font-style:italic;\"> </span>     . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">1.2  <span style=\" font-style:italic;\"> </span>    , , 5, 0, -24. </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">1.3  <span style=\" font-style:italic;\"> </span>   *, , 0.47, 0.535. </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">- <span style=\" font-style:italic;\">  </span>        .      . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-weight:600; font-style:italic;\">:</span>       -,       . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">1.5  <span style=\" font-style:italic;\"> </span>    *. </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">2. <span style=\" font-style:italic;\"> .</span>  ,     . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">2.2 <span style=\" font-style:italic;\">   </span> <span style=\" font-family:\'Symbol\';\">-</span>           . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">2.3 <span style=\" font-style:italic;\">    </span> <span style=\" font-family:\'Symbol\';\">-</span>           . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">2.4 <span style=\" font-style:italic;\">  </span> <span style=\" font-family:\'Symbol\';\">-</span>        ,     . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">3. <span style=\" font-style:italic;\"> </span>.  ,    ,      . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">3.1 <span style=\" font-style:italic;\">    </span> <span style=\" font-family:\'Symbol\';\">-</span>           . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">3.2 <span style=\" font-style:italic;\">  </span> <span style=\" font-family:\'Symbol\';\">-</span>         .        .   ,   ,    10%  ,      . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">3.3 <span style=\" font-style:italic;\">  </span> <span style=\" font-family:\'Symbol\';\">-</span>   ,     . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">3.4 <span style=\" font-style:italic;\">  </span> <span style=\" font-family:\'Symbol\';\">-</span>       ,         . </p>\n"
                                            "<p align=\"center\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-weight:600;\">2.2.  </span> </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"> <span style=\" font-style:italic;\"> </span>      ,         . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">                 . ,    10,           10. </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">  ,   ,    .       <span style=\" font-style:italic;\"> </span>   ,       .      <span style=\" font-style:italic;\"> </span>.     <span style=\" font-style:italic;\">   </span>    . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"> <span style=\" font-style:italic;\"> </span>       .            . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"> <span style=\" font-style:italic;\"> </span>        .               . ,       10,       10  .       10,    12,         10, 11  12 (). </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">        . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">    .. (), .. (),      .   .. (+),    .          ,    ,        <span style=\" font-style:italic;\"> </span>. </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">        .           / . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"> 3  4            *. </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"> 5      (  ,   ,    1). </p>\n"
                                            "<p align=\"center\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-weight:600;\">2.3.   </span> </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">     ,  <span style=\" font-style:italic;\"> </span>     . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">   ,    ,       .         . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">          5 . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">3       (  )    </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-weight:600; font-style:italic;\">!</span>  3  4       . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'Times New Roman\',\'serif\';\">   </span><span style=\" font-family:\'Times New Roman\',\'serif\'; font-style:italic;\">   </span><span style=\" font-family:\'Times New Roman\',\'serif\';\"> </span><span style=\" font-family:\'Times New Roman\',\'serif\'; font-style:italic;\">  </span><span style=\" font-family:\'Times New Roman\',\'serif\';\">          ,      (  )  .            .     -  .</span></p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">        <span style=\" font-style:italic;\"> </span>,          .         . </p>\n"
                                            "<p align=\"center\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-weight:600;\">2.4.  </span> </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">   11  (        )         .         . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"> 1           . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-weight:600; font-style:italic;\">:</span>      : M4_1, M4_2, V2_1, V2_2, G5_2, G5_3.       (M  , V   , G    ),     ,        . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"> 5       .        ,     <span style=\" font-style:italic;\">-</span>.         . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">  6  11       .     6-10         ,   11  ,      . </p>\n"
                                            "<p align=\"center\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-weight:600;\">2.5.  </span> </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">     ,  <span style=\" font-style:italic;\"> </span>     . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">   ,               .          . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">     . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">        <span style=\" font-style:italic;\"> </span>,          .         . </p>\n"
                                            "<p align=\"center\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-weight:600;\">2.6.  </span></p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"> 1     .    2- : <span style=\" font-style:italic;\"></span>   ,        <span style=\" font-style:italic;\"> </span>,    ;        <span style=\" font-style:italic;\">  </span>.             . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"> 11  12          .      .        . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">   <span style=\" font-style:italic;\">. .  </span>                     . </p>\n"
                                            "<p align=\"center\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-size:14pt; font-weight:600;\">3.  </span><span style=\" font-size:14pt;\"> </span></p>\n"
                                            "<p align=\"center\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-weight:600;\">3.1.   </span> </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">        . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">        <span style=\" font-style:italic;\"> </span>     , ,  1  v1.            <span style=\" font-style:italic;\">    _1</span> ( _v1). </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">             <span style=\" font-style:italic;\"> </span>.              <span style=\" font-style:italic;\"> </span>,       ,       . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">                  .            . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">         <span style=\" font-style:italic;\"> </span>.      ,            . </p>\n"
                                            "<p align=\"center\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-weight:600;\">3.2. </span> </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">        <span style=\" font-weight:600; font-style:italic;\">Excel-</span>  <span style=\" font-weight:600; font-style:italic;\">.xlsx</span>.      Excel- <span style=\" font-weight:600; font-style:italic;\">    </span>   ,       . </p>\n"
                                            "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">        .     .     ,      ,    1  v1         . </p></body></html>"))


#  
class Manual_Window(QtWidgets.QWidget, Ui_Form):
    def __init__(self):
        super().__init__()
        self.setupUi(self)


#       
class MyError(Exception):
    def __init__(self, error_text):
        self.txt = error_text

    def value_error_notify(self):
        path = os.path.join("SVO_Log.txt")
        logging.basicConfig(level=logging.DEBUG, filename=path,
                            format="%(asctime)s - %(module)s - %(levelname)s - %(funcName)s: %(lineno)d - %(message)s",
                            datefmt='%H:%M:%S')
        logging.exception('  !')
        MyError.error = QMessageBox()
        MyError.error.setWindowTitle("!")
        MyError.error.setIcon(QMessageBox.Warning)
        MyError.error.setStandardButtons(QMessageBox.Ok)
        MyError.error.setText('  !    .')
        path = os.path.join("data", "icon.png")
        icon = QtGui.QIcon(path)
        MyError.error.setWindowIcon(icon)
        MyError.error.exec_()
        return

    def atrribute_error_notify(self):
        path = os.path.join("SVO_Log.txt")
        logging.basicConfig(level=logging.DEBUG, filename=path,
                            format="%(asctime)s - %(module)s - %(levelname)s - %(funcName)s: %(lineno)d - %(message)s",
                            datefmt='%H:%M:%S')
        logging.exception('  !')
        MyError.error = QMessageBox()
        MyError.error.setWindowTitle("!")
        MyError.error.setIcon(QMessageBox.Warning)
        MyError.error.setStandardButtons(QMessageBox.Ok)
        MyError.error.setText('  !       .')
        path = os.path.join("data", "icon.png")
        icon = QtGui.QIcon(path)
        MyError.error.setWindowIcon(icon)
        MyError.error.exec_()
        return

    def save_success(self):
        MyError.error = QMessageBox()
        MyError.error.setWindowTitle(" ")
        MyError.error.setIcon(QMessageBox.Information)
        MyError.error.setStandardButtons(QMessageBox.Ok)
        MyError.error.setText('  .')
        path = os.path.join("data", "icon.png")
        icon = QtGui.QIcon(path)
        MyError.error.setWindowIcon(icon)
        MyError.error.exec_()
        self.pushButton_load_data.setDisabled(True)

    def save_fail(self):
        MyError.error = QMessageBox()
        MyError.error.setWindowTitle(" ")
        MyError.error.setIcon(QMessageBox.Warning)
        MyError.error.setStandardButtons(QMessageBox.Ok)
        MyError.error.setText('   .')
        path = os.path.join("data", "icon.png")
        icon = QtGui.QIcon(path)
        MyError.error.setWindowIcon(icon)
        MyError.error.exec_()


#  
class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.Manual = Manual_Window()
        self.pushButton_form_pr.setDisabled(True)
        self.comboBox_main_pr_type.setDisabled(True)
        self.pushButton_form_balance.setDisabled(True)
        self.pushButton_calc_balance.setDisabled(True)
        self.pushButton_2_room.clicked.connect(self.add_room)
        self.pushButton_add_row.clicked.connect(self.add_row)
        self.pushButton_3_del_row.clicked.connect(self.delete_rows)
        self.pushButton_4_calc_h_loss.clicked.connect(self.base_data)
        self.pushButton_4_calc_h_loss.clicked.connect(self.calc_heat_process)
        self.pushButton_4_calc_h_loss.clicked.connect(self.sum_loss_heat)
        self.pushButton_4_calc_h_loss.clicked.connect(self.sum_loss_heat_build)
        self.pushButton_form_pr.clicked.connect(self.get_data_from_h_loss)
        self.pushButton_calc_pr.clicked.connect(self.calc_pr_process)
        self.pushButton_calc_pr.clicked.connect(self.sum_heat_pr)
        self.pushButton_del_pr.clicked.connect(self.delete_pr_table)
        self.pushButton_add_section.clicked.connect(self.add_section)
        self.pushButton_del_section.clicked.connect(self.del_section)
        self.pushButton_hydro_calc.clicked.connect(self.calc_hydro_process)
        self.pushButton_hydro_calc.clicked.connect(self.sum_pressure_loss)
        self.pushButton_form_balance.clicked.connect(self.get_data_from_hydro)
        self.pushButton_del_balance.clicked.connect(self.delete_balance_table)
        self.pushButton_calc_balance.clicked.connect(self.calc_balance_process)
        self.pushButton_add_col.clicked.connect(self.tube_add_column)
        self.pushButton_del_col.clicked.connect(self.tube_del_column)
        self.pushButton_tube_calc.clicked.connect(self.tube_calc_process)
        self.pushButton_del_files.clicked.connect(self.delete_files)
        self.pushButton_save_data.clicked.connect(self.save_main_data)
        self.pushButton_save_data.clicked.connect(self.save_heat_data)
        self.pushButton_save_data.clicked.connect(self.save_pr_data)
        self.pushButton_save_data.clicked.connect(self.save_hydro_data)
        self.pushButton_save_data.clicked.connect(self.save_balance_data)
        self.pushButton_save_data.clicked.connect(self.save_tube_data)
        self.pushButton_save_data.clicked.connect(self.json_save_data)
        self.pushButton_load_data.clicked.connect(self.json_load_data)
        self.pushButton_load_data.clicked.connect(self.load_main_data)
        self.pushButton_load_data.clicked.connect(self.load_heat_data)
        self.pushButton_load_data.clicked.connect(self.load_pr_data)
        self.pushButton_load_data.clicked.connect(self.load_hydro_data)
        self.pushButton_load_data.clicked.connect(self.load_balance_data)
        self.pushButton_load_data.clicked.connect(self.load_tube_data)
        self.pushButton_import.clicked.connect(self.heat_data_to_excel)
        self.pushButton_import.clicked.connect(self.pr_data_to_excel)
        self.pushButton_import.clicked.connect(self.hydro_data_to_excel)
        self.pushButton_import.clicked.connect(self.balance_data_to_excel)
        self.pushButton_import.clicked.connect(self.tube_data_to_excel)
        self.pushButton_import.clicked.connect(self.import_data_to_excel)
        self.pushButton_7_help.clicked.connect(self.manual_ex)
        self.checkBox_3_n_coef.stateChanged.connect(self.norm_coef)
        self.comboBox_main_pr_type.activated.connect(self.main_type_pr_choose)
        self.comboBox_main_connect_type.activated.connect(self.connect_type_choose)
        self.comboBox_wind_style.activated.connect(self.change_style)
        self.comboBox_main_pr_type.setCurrentIndex(-1)
        self.comboBox_wind_style.setCurrentIndex(-1)
        self.comboBox_q_r_choose.activated.connect(self.tube_main_calc_way)
        global room_number
        room_number = 1
        self.file_list = []
        self.ns_room_long = []
        path = os.path.join("data", "SVO_data.json")
        my_file = open(path, "a")
        my_file.close()
        try:
            with open(path, "r") as read_file:
                self.json_data = json.load(read_file)
                self.file_list_data = self.json_data["file_list"]
                self.file_list.extend(self.file_list_data)
                for i in range(len(self.file_list)):
                    self.comboBox_load_from.addItem(str(self.file_list[i]))
        except:
            pass
        self.t1 = int(self.comboBox_t1.currentText())
        self.t2 = int(self.comboBox_2_t2.currentText())

    def add_room(self):
        global room_number

        self.num_symb = QtWidgets.QTableWidgetItem('')
        self.num_symb.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
        self.num_symb.setBackground(QtGui.QColor(85, 255, 255))
        self.num_symb.setFlags(QtCore.Qt.ItemIsEnabled)

        self.room_order = QtWidgets.QTableWidgetItem(str(room_number))
        self.room_order.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
        self.room_order.setBackground(QtGui.QColor(170, 255, 127))

        self.room_name = QtWidgets.QTableWidgetItem('')
        self.room_name.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
        self.room_name.setBackground(QtGui.QColor(170, 255, 127))

        self.t_vn_name = QtWidgets.QTableWidgetItem('t, C')
        self.t_vn_name.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
        self.t_vn_name.setBackground(QtGui.QColor(85, 255, 255))
        self.t_vn_name.setFlags(QtCore.Qt.ItemIsEnabled)

        self.t_vn = QtWidgets.QTableWidgetItem('')
        self.t_vn.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
        self.t_vn.setBackground(QtGui.QColor(170, 255, 127))

        self.sum_room_heat_loss = QtWidgets.QTableWidgetItem('')
        self.sum_room_heat_loss.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
        self.sum_room_heat_loss.setBackground(QtGui.QColor(85, 255, 255))
        self.sum_room_heat_loss.setFlags(QtCore.Qt.ItemIsEnabled)

        rowPosition = self.tableWidget.rowCount()
        if self.lineEdit_4_room.text() == "":
            self.tableWidget.insertRow(rowPosition)
            self.tableWidget.setItem(rowPosition, 0, self.num_symb)
            self.tableWidget.setItem(rowPosition, 1, self.room_order)
            self.tableWidget.setItem(rowPosition, 2, self.room_name)
            self.tableWidget.setItem(rowPosition, 3, self.t_vn_name)
            self.tableWidget.setItem(rowPosition, 4, self.t_vn)
            self.tableWidget.setItem(rowPosition, 5, self.sum_room_heat_loss)
            room_number += 1
        else:
            try:
                int_room = int(self.lineEdit_4_room.text())
            except ValueError:
                MyError.value_error_notify(self)

            self.tableWidget.insertRow(int_room)
            self.tableWidget.setItem(int_room, 0, self.num_symb)
            self.tableWidget.setItem(int_room, 1, self.room_order)
            self.tableWidget.setItem(int_room, 2, self.room_name)
            self.tableWidget.setItem(int_room, 3, self.t_vn_name)
            self.tableWidget.setItem(int_room, 4, self.t_vn)
            self.tableWidget.setItem(int_room, 5, self.sum_room_heat_loss)
            self.lineEdit_4_room.clear()
            room_number += 1
        self.tableWidget.resizeRowsToContents()
        self.tableWidget.resizeColumnsToContents()

    def add_row(self):
        self.combo_f_structure = QtWidgets.QComboBox()
        self.combo_f_structure.addItem(' I .')
        self.combo_f_structure.addItem(' II .')
        self.combo_f_structure.addItem(' III .')
        self.combo_f_structure.addItem(' IV .')
        self.combo_f_structure.addItem('')
        self.combo_f_structure.addItem('..')
        self.combo_f_structure.addItem('.. ()')
        self.combo_f_structure.addItem('.. ()')
        self.combo_f_structure.addItem('.. (+)')
        self.combo_f_structure.addItem(self.lineEdit_another_ok.text())
        self.combo_f_structure.currentIndexChanged.connect(self.combo_cond)

        self.combo_s_world = QtWidgets.QComboBox()
        self.combo_s_world.addItem('')
        self.combo_s_world.addItem('')
        self.combo_s_world.addItem('')
        self.combo_s_world.addItem('')
        self.combo_s_world.addItem('')
        self.combo_s_world.addItem('')
        self.combo_s_world.addItem('')
        self.combo_s_world.addItem('')
        self.combo_s_world.addItem('')
        self.combo_s_world.setCurrentIndex(8)
        self.combo_s_world.currentIndexChanged.connect(self.side_world)

        self.room_len = QtWidgets.QTableWidgetItem('')
        self.room_len.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)

        self.room_wid = QtWidgets.QTableWidgetItem('')
        self.room_wid.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)

        self.block_quan = QtWidgets.QTableWidgetItem('1')
        self.block_quan.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)

        self.sum_heat_loss = QtWidgets.QTableWidgetItem('')
        self.sum_heat_loss.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
        self.sum_heat_loss.setFlags(QtCore.Qt.ItemIsEnabled)

        rowPosition = self.tableWidget.rowCount()
        if self.lineEdit_3_add_row.text() == "":
            self.tableWidget.insertRow(rowPosition)
            self.tableWidget.setCellWidget(rowPosition, 0, self.combo_f_structure)
            self.tableWidget.setCellWidget(rowPosition, 1, self.combo_s_world)
            self.tableWidget.setItem(rowPosition, 2, self.room_len)
            self.tableWidget.setItem(rowPosition, 3, self.room_wid)
            self.tableWidget.setItem(rowPosition, 4, self.block_quan)
            self.tableWidget.setItem(rowPosition, 5, self.sum_heat_loss)
        else:
            try:
                int_row = int(self.lineEdit_3_add_row.text())
            except ValueError:
                MyError.value_error_notify(self)
            self.tableWidget.insertRow(int_row)
            self.tableWidget.setCellWidget(int_row, 0, self.combo_f_structure)
            self.tableWidget.setCellWidget(int_row, 1, self.combo_s_world)
            self.tableWidget.setItem(int_row, 2, self.room_len)
            self.tableWidget.setItem(int_row, 3, self.room_wid)
            self.tableWidget.setItem(int_row, 4, self.block_quan)
            self.tableWidget.setItem(int_row, 5, self.sum_heat_loss)
            self.lineEdit_3_add_row.clear()
        self.tableWidget.resizeRowsToContents()
        self.tableWidget.resizeColumnsToContents()

    def combo_cond(self):
        rowPosition = self.tableWidget.currentRow() + 1

        self.room_len = QtWidgets.QTableWidgetItem('')
        self.room_len.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)

        self.room_len_1 = QtWidgets.QTableWidgetItem('')
        self.room_len_1.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)

        self.room_wid = QtWidgets.QTableWidgetItem('')
        self.room_wid.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)

        self.room_wid_1 = QtWidgets.QTableWidgetItem('')
        self.room_wid_1.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)

        self.block_quan = QtWidgets.QTableWidgetItem('1')
        self.block_quan.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)

        self.block_quan_1 = QtWidgets.QTableWidgetItem('1')
        self.block_quan_1.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)

        self.sum_heat_loss = QtWidgets.QTableWidgetItem('')
        self.sum_heat_loss.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
        self.sum_heat_loss.setFlags(QtCore.Qt.ItemIsEnabled)

        self.sum_heat_loss_1 = QtWidgets.QTableWidgetItem('')
        self.sum_heat_loss_1.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
        self.sum_heat_loss_1.setFlags(QtCore.Qt.ItemIsEnabled)

        self.room_window = QtWidgets.QTableWidgetItem('')
        self.room_window.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)

        self.room_door = QtWidgets.QTableWidgetItem('')
        self.room_door.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)

        self.side_world_1 = QtWidgets.QTableWidgetItem('')
        self.side_world_1.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
        self.side_world_1.setFlags(QtCore.Qt.ItemIsEnabled)

        self.side_world_2 = QtWidgets.QTableWidgetItem('')
        self.side_world_2.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
        self.side_world_2.setFlags(QtCore.Qt.ItemIsEnabled)

        self.current_item = self.tableWidget.cellWidget(self.tableWidget.currentRow(), self.tableWidget.currentColumn())
        self.side_item = self.tableWidget.cellWidget(self.tableWidget.currentRow(),
                                                     self.tableWidget.currentColumn() + 1)
        if isinstance(self.current_item, QtWidgets.QComboBox):
            cur_text = self.current_item.currentIndex()
            for i in range(5):
                if cur_text == i:
                    self.side_item.setCurrentIndex(8)

            if cur_text == 9:
                self.side_item.setCurrentIndex(8)

            if cur_text == 6:
                self.tableWidget.insertRow(rowPosition)
                self.tableWidget.setItem(rowPosition, 0, self.room_window)
                self.tableWidget.setItem(rowPosition, 1, self.side_world_1)
                self.tableWidget.item(rowPosition, 1).setText(
                    self.tableWidget.cellWidget(rowPosition - 1, 1).currentText())
                self.tableWidget.setItem(rowPosition, 2, self.room_len)
                self.tableWidget.setItem(rowPosition, 3, self.room_wid)
                self.tableWidget.setItem(rowPosition, 4, self.block_quan)
                self.tableWidget.setItem(rowPosition, 5, self.sum_heat_loss)

            if cur_text == 7:
                self.tableWidget.insertRow(rowPosition)
                self.tableWidget.setItem(rowPosition, 0, self.room_door)
                self.tableWidget.setItem(rowPosition, 1, self.side_world_1)
                self.tableWidget.item(rowPosition, 1).setText(
                    self.tableWidget.cellWidget(rowPosition - 1, 1).currentText())
                self.tableWidget.setItem(rowPosition, 2, self.room_len)
                self.tableWidget.setItem(rowPosition, 3, self.room_wid)
                self.tableWidget.setItem(rowPosition, 4, self.block_quan)
                self.tableWidget.setItem(rowPosition, 5, self.sum_heat_loss)

            if cur_text == 8:
                self.tableWidget.insertRow(rowPosition)
                self.tableWidget.setItem(rowPosition, 0, self.room_window)
                self.tableWidget.setItem(rowPosition, 1, self.side_world_1)
                self.tableWidget.item(rowPosition, 1).setText(
                    self.tableWidget.cellWidget(rowPosition - 1, 1).currentText())
                self.tableWidget.setItem(rowPosition, 2, self.room_len)
                self.tableWidget.setItem(rowPosition, 3, self.room_wid)
                self.tableWidget.setItem(rowPosition, 4, self.block_quan)
                self.tableWidget.setItem(rowPosition, 5, self.sum_heat_loss)
                self.tableWidget.insertRow(rowPosition + 1)
                self.tableWidget.setItem(rowPosition + 1, 0, self.room_door)
                self.tableWidget.setItem(rowPosition + 1, 1, self.side_world_2)
                self.tableWidget.item(rowPosition + 1, 1).setText(
                    self.tableWidget.cellWidget(rowPosition - 1, 1).currentText())
                self.tableWidget.setItem(rowPosition + 1, 2, self.room_len_1)
                self.tableWidget.setItem(rowPosition + 1, 3, self.room_wid_1)
                self.tableWidget.setItem(rowPosition + 1, 4, self.block_quan_1)
                self.tableWidget.setItem(rowPosition + 1, 5, self.sum_heat_loss_1)

    def side_world(self):
        current_item = self.tableWidget.cellWidget(self.tableWidget.currentRow(), self.tableWidget.currentColumn())
        side_item_1 = self.tableWidget.item(self.tableWidget.currentRow() + 1, self.tableWidget.currentColumn())
        side_item_2 = self.tableWidget.item(self.tableWidget.currentRow() + 2, self.tableWidget.currentColumn())
        if isinstance(side_item_1, QtWidgets.QTableWidgetItem):
            side_item_1.setText(current_item.currentText())
            if isinstance(side_item_2, QtWidgets.QTableWidgetItem):
                side_item_2.setText(current_item.currentText())

    def delete_rows(self):
        try:
            rowPosition = self.tableWidget.rowCount() - 1
            if self.lineEdit_f_d_row.text() == "" or self.lineEdit_2_l_d_row.text() == "":
                self.tableWidget.removeRow(rowPosition)
            else:
                f_int_row = int(self.lineEdit_f_d_row.text()) - 1
                l_int_row = int(self.lineEdit_2_l_d_row.text())
                for i in range(f_int_row, l_int_row):
                    self.tableWidget.removeRow(f_int_row)
                self.lineEdit_f_d_row.clear()
                self.lineEdit_2_l_d_row.clear()
        except ValueError:
            MyError.value_error_notify(self)

    def norm_coef(self):
        if self.checkBox_3_n_coef.isChecked():
            self.lineEdit_9_ns_20.setText(str(0.47))
            self.lineEdit_9_ns_16.setText(str(0.5))
            self.lineEdit_gate.setText(str(2.5))
            self.lineEdit_12_wind.setText(str(1.59))
            self.lineEdit_16_door.setText(str(2.33))
            self.lineEdit_15_kr.setText(str(0.32))
            self.lineEdit_11_pol_1.setText(str(0.48))
            self.lineEdit_13_pol_2.setText(str(0.23))
            self.lineEdit_7_pol_3.setText(str(0.12))
            self.lineEdit_14_pol_4.setText(str(0.07))
        else:
            self.lineEdit_9_ns_20.clear()
            self.lineEdit_9_ns_16.clear()
            self.lineEdit_gate.clear()
            self.lineEdit_12_wind.clear()
            self.lineEdit_16_door.clear()
            self.lineEdit_15_kr.clear()
            self.lineEdit_11_pol_1.clear()
            self.lineEdit_13_pol_2.clear()
            self.lineEdit_7_pol_3.clear()
            self.lineEdit_14_pol_4.clear()

    def base_data(self):
        try:
            self.t_v_out = int(self.lineEdit_6_t_n.text())
            self.t1 = int(self.comboBox_t1.currentText())
            self.t2 = int(self.comboBox_2_t2.currentText())
            self.p1 = float(self.lineEdit_p_n.text())
            self.k_mass = (
                float(self.lineEdit_11_pol_1.text()), float(self.lineEdit_13_pol_2.text()),
                float(self.lineEdit_7_pol_3.text()), float(self.lineEdit_14_pol_4.text()),
                float(self.lineEdit_15_kr.text()), float(self.lineEdit_9_ns_16.text()),
                float(self.lineEdit_9_ns_20.text()), float(self.lineEdit_12_wind.text()),
                float(self.lineEdit_16_door.text()), float(self.lineEdit_gate.text()))
            self.pushButton_form_pr.setDisabled(False)
        except ValueError:
            self.pushButton_form_pr.setDisabled(True)
            MyError.value_error_notify(self)

    def calc_heat_process(self):
        try:
            for i in range(self.tableWidget.rowCount()):
                coef_str = 0
                coef_square = 0
                self.current_item = self.tableWidget.item(i, 0)
                self.current_side = self.tableWidget.item(i, 1)
                if isinstance(self.current_item, QtWidgets.QTableWidgetItem):
                    self.current_item = self.tableWidget.item(i, 0)
                    self.current_text = self.tableWidget.item(i, 0).text()
                    if self.current_text == '':
                        self.k_tp = self.k_mass[7]
                    if self.current_text == '':
                        self.k_tp = self.k_mass[8]
                    if self.current_text == '':
                        if int(self.tableWidget.item(i, 4).text()) >= 16:
                            k_tp = self.k_mass[6]
                        else:
                            k_tp = self.k_mass[5]
                else:
                    self.current_item = self.tableWidget.cellWidget(i, 0)
                    self.current_text = self.tableWidget.cellWidget(i, 0).currentText()
                    for m in range(4):
                        if self.current_item.currentIndex() == m:
                            coef_str = 1
                    for m in range(5):
                        if self.current_item.currentIndex() == m:
                            self.k_tp = self.k_mass[m]
                    for m in range(5, 9):
                        if self.current_item.currentIndex() == m:
                            self.k_tp = k_tp
                    if self.current_item.currentIndex() == 9:
                        self.k_tp = self.k_mass[9]
                        coef_str = 1
                    if self.current_item.currentIndex() == 6 or self.current_item.currentIndex() == 7:
                        coef_square = 1
                    if self.current_item.currentIndex() == 8:
                        coef_square = 2

                if isinstance(self.current_side, QtWidgets.QTableWidgetItem):
                    self.current_side = self.tableWidget.item(i, 1)
                    self.current_side_text = self.tableWidget.item(i, 1).text()
                else:
                    self.current_side = self.tableWidget.cellWidget(i, 1)
                    self.current_side_text = self.tableWidget.cellWidget(i, 1).currentText()

                if self.current_text == '':
                    self.delta_t = int(self.tableWidget.item(i, 4).text()) - self.t_v_out
                else:
                    if coef_str == 1:
                        coef_struct = 1
                    elif self.current_side_text == '' or self.current_side_text == '' or self.current_side_text == 'C' or self.current_side_text == 'C':
                        coef_struct = 1.2
                    elif self.current_side_text == '' or self.current_side_text == '' or self.current_side_text == '' or self.current_side_text == '':
                        coef_struct = 1.15
                    else:
                        coef_struct = 1.1
                    if coef_square == 1:
                        self.square = float(self.tableWidget.item(i, 2).text()) * float(
                            self.tableWidget.item(i, 3).text()) * float(self.tableWidget.item(i, 4).text()) - float(
                            self.tableWidget.item(i + 1, 2).text()) * float(
                            self.tableWidget.item(i + 1, 3).text()) * float(self.tableWidget.item(i + 1, 4).text())
                    elif coef_square == 2:
                        self.square = float(self.tableWidget.item(i, 2).text()) * float(
                            self.tableWidget.item(i, 3).text()) * float(self.tableWidget.item(i, 4).text()) - float(
                            self.tableWidget.item(i + 1, 2).text()) * float(
                            self.tableWidget.item(i + 1, 3).text()) * float(
                            self.tableWidget.item(i + 1, 4).text()) - float(
                            self.tableWidget.item(i + 2, 2).text()) * float(
                            self.tableWidget.item(i + 2, 3).text()) * float(self.tableWidget.item(i + 2, 4).text())
                    else:
                        self.square = float(self.tableWidget.item(i, 2).text()) * float(
                            self.tableWidget.item(i, 3).text()) * float(self.tableWidget.item(i, 4).text())
                    self.heat_loss = round(self.k_tp * self.square * self.delta_t * coef_struct, 0)
                    self.tableWidget.item(i, 5).setText(str(int(self.heat_loss)))
        except ValueError:
            self.pushButton_form_pr.setDisabled(True)
            MyError.value_error_notify(self)
        except AttributeError:
            MyError.value_error_notify(self)

    def sum_loss_heat(self):
        self.pushButton_form_pr.setDisabled(False)
        try:
            self.sum_loss_h = 0
            self.sum_ns_long = 0
            if self.checkBox_inf.isChecked():
                coef_inf = 1.2
            else:
                coef_inf = 1
            if self.checkBox_2_t_reg.isChecked():
                coef_sp_60 = 1.15
            else:
                coef_sp_60 = 1
            for i in range(self.tableWidget.rowCount()):
                if i + 1 > self.tableWidget.rowCount():
                    break
                if self.tableWidget.item(i, 3).text() == 't, C':
                    a = i
                elif self.tableWidget.item(i, 3).text() != 't, C':
                    self.sum_loss_h += int(self.tableWidget.item(i, 5).text())
                    try:
                        if self.tableWidget.cellWidget(i, 0).currentIndex() == 5 or self.tableWidget.cellWidget(i,
                                                                                                                0).currentIndex() == 6 or self.tableWidget.cellWidget(
                                i, 0).currentIndex() == 7 or self.tableWidget.cellWidget(i, 0).currentIndex() == 8:
                            self.sum_ns_long += float(self.tableWidget.item(i, 2).text())
                    except:
                        pass
                    if i + 1 == self.tableWidget.rowCount() or self.tableWidget.item(i + 1, 3).text() == 't, C':
                        self.tableWidget.item(a, 5).setText(str(int(round(coef_inf * coef_sp_60 * self.sum_loss_h, 0))))
                        self.ns_room_long.append(self.sum_ns_long)
                        self.sum_loss_h = 0
                        self.sum_ns_long = 0
        except ValueError:
            return

    def sum_loss_heat_build(self):
        try:
            loss_heat_build = 0
            for i in range(self.tableWidget.rowCount()):
                if self.tableWidget.item(i, 3).text() == 't, C':
                    loss_heat_build += int(self.tableWidget.item(i, 5).text())
                if i + 1 == self.tableWidget.rowCount():
                    bal_g_so = float(round(3600 * loss_heat_build / (4200 * (self.t1 - self.t2)), 1))
                    self.lineEdit_17_sum_h_loss.setText(str(float(round(loss_heat_build / 1000, 2))))
                    self.lineEdit_24_n_rate.setText(str(bal_g_so))
        except ValueError:
            return

    def get_data_from_h_loss(self):
        rowPosition = 0
        for i in range(self.tableWidget.rowCount()):
            if self.tableWidget.item(i, 3).text() == 't, C':
                self.tableWidget_2.insertRow(rowPosition)

                self.pr_type_name = QtWidgets.QTableWidgetItem(' ')
                self.pr_type_name.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.pr_type_name.setBackground(QtGui.QColor(85, 255, 255))
                self.pr_type_name.setFlags(QtCore.Qt.ItemIsEnabled)

                self.pr_type = QtWidgets.QComboBox()
                self.pr_type.setStyleSheet("background-color: rgb(170, 255, 127);")
                self.tableWidget_2.setCellWidget(rowPosition, 4, self.pr_type)
                self.tableWidget_2.cellWidget(rowPosition, 4).addItem('')
                self.tableWidget_2.cellWidget(rowPosition, 4).addItem('')
                self.tableWidget_2.cellWidget(rowPosition, 4).addItem('  . .')
                self.tableWidget_2.cellWidget(rowPosition, 4).setCurrentIndex(-1)
                self.pr_type.activated.connect(self.local_type_pr_choose)

                self.num_symb = QtWidgets.QTableWidgetItem('')
                self.num_symb.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.num_symb.setBackground(QtGui.QColor(85, 255, 255))
                self.num_symb.setFlags(QtCore.Qt.ItemIsEnabled)

                self.room_order = QtWidgets.QTableWidgetItem(str(room_number))
                self.room_order.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.room_order.setBackground(QtGui.QColor(85, 255, 255))
                self.room_order.setFlags(QtCore.Qt.ItemIsEnabled)

                self.room_name = QtWidgets.QTableWidgetItem('')
                self.room_name.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.room_name.setBackground(QtGui.QColor(85, 255, 255))
                self.room_name.setFlags(QtCore.Qt.ItemIsEnabled)

                self.tableWidget_2.setItem(rowPosition, 0, self.num_symb)
                self.tableWidget_2.setItem(rowPosition, 1, self.room_order)

                self.tableWidget_2.item(rowPosition, 1).setText(self.tableWidget.item(i, 1).text())
                self.tableWidget_2.setItem(rowPosition, 2, self.room_name)

                self.tableWidget_2.item(rowPosition, 2).setText(self.tableWidget.item(i, 2).text())
                self.tableWidget_2.setItem(rowPosition, 3, self.pr_type_name)

                rowPosition += 1
                self.tableWidget_2.insertRow(rowPosition)

                self.size_type = QtWidgets.QTableWidgetItem('')
                self.size_type.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.size_type.setFlags(QtCore.Qt.ItemIsEnabled)

                self.line_num_name = QtWidgets.QTableWidgetItem('')
                self.line_num_name.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.line_num_name.setFlags(QtCore.Qt.ItemIsEnabled)

                self.q_n_name = QtWidgets.QTableWidgetItem('')
                self.q_n_name.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.q_n_name.setFlags(QtCore.Qt.ItemIsEnabled)

                self.connect_type_name = QtWidgets.QTableWidgetItem(' ')
                self.connect_type_name.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.connect_type_name.setFlags(QtCore.Qt.ItemIsEnabled)

                self.pr_quan_name = QtWidgets.QTableWidgetItem('- , ')
                self.pr_quan_name.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.pr_quan_name.setFlags(QtCore.Qt.ItemIsEnabled)

                self.tableWidget_2.setItem(rowPosition, 0, self.size_type)
                self.tableWidget_2.setItem(rowPosition, 1, self.line_num_name)
                self.tableWidget_2.setItem(rowPosition, 2, self.q_n_name)
                self.tableWidget_2.setItem(rowPosition, 3, self.connect_type_name)
                self.tableWidget_2.setItem(rowPosition, 4, self.pr_quan_name)

                rowPosition += 1
                self.tableWidget_2.insertRow(rowPosition)

                self.pr_quan = QtWidgets.QTableWidgetItem('1')
                self.pr_quan.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.pr_quan.setBackground(QtGui.QColor(170, 255, 127))
                self.tableWidget_2.setItem(rowPosition, 4, self.pr_quan)

                rowPosition += 1
                self.tableWidget_2.insertRow(rowPosition)

                self.size_gab_name = QtWidgets.QTableWidgetItem('')
                self.size_gab_name.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.size_gab_name.setFlags(QtCore.Qt.ItemIsEnabled)

                self.q_r_pr_name = QtWidgets.QTableWidgetItem('Q, ')
                self.q_r_pr_name.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.q_r_pr_name.setFlags(QtCore.Qt.ItemIsEnabled)

                self.q_n_pr_name = QtWidgets.QTableWidgetItem('Q, ')
                self.q_n_pr_name.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.q_n_pr_name.setFlags(QtCore.Qt.ItemIsEnabled)

                self.g_pr_name = QtWidgets.QTableWidgetItem('G, /')
                self.g_pr_name.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.g_pr_name.setFlags(QtCore.Qt.ItemIsEnabled)

                self.q_r_room_name = QtWidgets.QTableWidgetItem('Q_, ')
                self.q_r_room_name.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.q_r_room_name.setFlags(QtCore.Qt.ItemIsEnabled)

                self.tableWidget_2.setItem(rowPosition, 0, self.size_gab_name)
                self.tableWidget_2.setItem(rowPosition, 1, self.q_r_pr_name)
                self.tableWidget_2.setItem(rowPosition, 2, self.q_n_pr_name)
                self.tableWidget_2.setItem(rowPosition, 3, self.g_pr_name)
                self.tableWidget_2.setItem(rowPosition, 4, self.q_r_room_name)

                rowPosition += 1
                self.tableWidget_2.insertRow(rowPosition)

                self.size_gab = QtWidgets.QTableWidgetItem('')
                self.size_gab.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.size_gab.setFlags(QtCore.Qt.ItemIsEnabled)

                self.q_r_pr = QtWidgets.QTableWidgetItem('')
                self.q_r_pr.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.q_r_pr.setFlags(QtCore.Qt.ItemIsEnabled)

                self.q_n_pr = QtWidgets.QTableWidgetItem('')
                self.q_n_pr.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.q_n_pr.setFlags(QtCore.Qt.ItemIsEnabled)

                self.g_pr = QtWidgets.QTableWidgetItem('')
                self.g_pr.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.g_pr.setFlags(QtCore.Qt.ItemIsEnabled)

                self.q_r_room = QtWidgets.QTableWidgetItem('')
                self.q_r_room.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.q_r_room.setFlags(QtCore.Qt.ItemIsEnabled)

                self.tableWidget_2.setItem(rowPosition, 0, self.size_gab)
                self.tableWidget_2.setItem(rowPosition, 1, self.q_r_pr)
                self.tableWidget_2.setItem(rowPosition, 2, self.q_n_pr)
                self.tableWidget_2.setItem(rowPosition, 3, self.g_pr)
                self.tableWidget_2.setItem(rowPosition, 4, self.q_r_room)

                rowPosition += 1

        self.tableWidget.resizeRowsToContents()
        self.tableWidget.resizeColumnsToContents()
        self.pushButton_add_row.setDisabled(True)
        self.pushButton_3_del_row.setDisabled(True)
        self.pushButton_2_room.setDisabled(True)
        self.pushButton_4_calc_h_loss.setDisabled(True)
        self.comboBox_main_pr_type.setDisabled(False)
        self.pushButton_form_pr.setDisabled(True)

    def type_pr_choose(self):
        if self.tableWidget_2.cellWidget(self.rowPosition, 4).currentIndex() == 0:
            self.tableWidget_2.item(self.rowPosition + 1, 0).setText(' , ')
            self.tableWidget_2.item(self.rowPosition + 1, 1).setText('')
            self.tableWidget_2.item(self.rowPosition + 1, 2).setText('Q/, ')

            self.height_pr_sec = QtWidgets.QComboBox()
            self.height_pr_sec.addItem('500')
            self.height_pr_sec.addItem('350')
            self.height_pr_sec.setStyleSheet("background-color: rgb(170, 255, 127);")
            self.tableWidget_2.setCellWidget(self.rowPosition + 2, 0, self.height_pr_sec)
            self.height_pr_sec.currentIndexChanged.connect(self.sec_height_choose)

            self.line_check = self.tableWidget_2.cellWidget(self.rowPosition + 2, 1)
            if isinstance(self.line_check, QtWidgets.QComboBox):
                self.tableWidget_2.removeCellWidget(self.rowPosition + 2, 1)
            self.line_num = QtWidgets.QTableWidgetItem('')
            self.line_num.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.line_num.setFlags(QtCore.Qt.ItemIsEnabled)
            self.tableWidget_2.setItem(self.rowPosition + 2, 1, self.line_num)

            self.q_n = QtWidgets.QComboBox()
            self.q_n.addItem('160')
            self.q_n.addItem('170')
            self.q_n.addItem('180')
            self.q_n.setStyleSheet("background-color: rgb(170, 255, 127);")
            self.tableWidget_2.setCellWidget(self.rowPosition + 2, 2, self.q_n)
            self.tableWidget_2.cellWidget(self.rowPosition + 2, 2).setCurrentIndex(2)

            self.connect_type = QtWidgets.QComboBox()
            self.connect_type.addItem('')
            self.connect_type.addItem(' -')
            self.connect_type.addItem(' -')
            self.connect_type.addItem(' -')
            self.connect_type.addItem(' -')
            self.connect_type.addItem(' -')
            self.connect_type.setStyleSheet("background-color: rgb(170, 255, 127);")
            self.tableWidget_2.setCellWidget(self.rowPosition + 2, 3, self.connect_type)
            self.tableWidget_2.item(self.rowPosition + 3, 0).setText('N, ')

        if self.tableWidget_2.cellWidget(self.rowPosition, 4).currentIndex() == 1:
            self.tableWidget_2.item(self.rowPosition + 1, 0).setText(' , ')
            self.tableWidget_2.item(self.rowPosition + 1, 1).setText(' ')
            self.tableWidget_2.item(self.rowPosition + 1, 2).setText('Q, ')

            self.height_pr_pl = QtWidgets.QComboBox()
            self.height_pr_pl.addItem('300')
            self.height_pr_pl.addItem('400')
            self.height_pr_pl.addItem('500')
            self.height_pr_pl.addItem('600')
            self.height_pr_pl.setStyleSheet("background-color: rgb(170, 255, 127);")
            self.tableWidget_2.setCellWidget(self.rowPosition + 2, 0, self.height_pr_pl)

            self.line_num = QtWidgets.QComboBox()
            self.line_num.addItem('1')
            self.line_num.addItem('2')
            self.line_num.setStyleSheet("background-color: rgb(170, 255, 127);")
            self.tableWidget_2.setCellWidget(self.rowPosition + 2, 1, self.line_num)

            self.q_n_check = self.tableWidget_2.cellWidget(self.rowPosition + 2, 2)
            if isinstance(self.q_n_check, QtWidgets.QComboBox):
                self.tableWidget_2.removeCellWidget(self.rowPosition + 2, 2)
            self.q_n = QtWidgets.QTableWidgetItem('')
            self.q_n.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.q_n.setFlags(QtCore.Qt.ItemIsEnabled)
            self.tableWidget_2.setItem(self.rowPosition + 2, 2, self.q_n)

            self.connect_type = QtWidgets.QComboBox()
            self.connect_type.addItem('')
            self.connect_type.addItem(' -')
            self.connect_type.addItem(' -')
            self.connect_type.addItem(' -')
            self.connect_type.addItem(' -')
            self.connect_type.addItem(' -')
            self.connect_type.setStyleSheet("background-color: rgb(170, 255, 127);")
            self.tableWidget_2.setCellWidget(self.rowPosition + 2, 3, self.connect_type)
            self.tableWidget_2.item(self.rowPosition + 3, 0).setText('L, ')

        if self.tableWidget_2.cellWidget(self.rowPosition, 4).currentIndex() == 2:
            self.tableWidget_2.item(self.rowPosition + 1, 0).setText(' , ')
            self.tableWidget_2.item(self.rowPosition + 1, 1).setText(' ')
            self.tableWidget_2.item(self.rowPosition + 1, 2).setText('Q, ')

            self.d_pr_reg = QtWidgets.QComboBox()
            self.d_pr_reg.addItem('32')
            self.d_pr_reg.addItem('40')
            self.d_pr_reg.addItem('50')
            self.d_pr_reg.addItem('65')
            self.d_pr_reg.addItem('80')
            self.d_pr_reg.addItem('100')
            self.d_pr_reg.addItem('125')
            self.d_pr_reg.addItem('150')
            self.d_pr_reg.setStyleSheet("background-color: rgb(170, 255, 127);")
            self.tableWidget_2.setCellWidget(self.rowPosition + 2, 0, self.d_pr_reg)

            self.line_num = QtWidgets.QComboBox()
            self.line_num.addItem('1')
            self.line_num.addItem('2')
            self.line_num.addItem('3')
            self.line_num.addItem('4')
            self.line_num.addItem('5')
            self.line_num.addItem('6')
            self.line_num.setStyleSheet("background-color: rgb(170, 255, 127);")
            self.tableWidget_2.setCellWidget(self.rowPosition + 2, 1, self.line_num)

            self.q_n_check = self.tableWidget_2.cellWidget(self.rowPosition + 2, 2)
            if isinstance(self.q_n_check, QtWidgets.QComboBox):
                self.tableWidget_2.removeCellWidget(self.rowPosition + 2, 2)
            self.q_n = QtWidgets.QTableWidgetItem('')
            self.q_n.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.q_n.setFlags(QtCore.Qt.ItemIsEnabled)
            self.tableWidget_2.setItem(self.rowPosition + 2, 2, self.q_n)

            self.connect_type_check = self.tableWidget_2.cellWidget(self.rowPosition + 2, 3)
            if isinstance(self.connect_type_check, QtWidgets.QComboBox):
                self.tableWidget_2.removeCellWidget(self.rowPosition + 2, 3)
            self.connect_type = QtWidgets.QTableWidgetItem('')
            self.connect_type.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.connect_type.setFlags(QtCore.Qt.ItemIsEnabled)
            self.tableWidget_2.setItem(self.rowPosition + 2, 3, self.connect_type)
            self.tableWidget_2.item(self.rowPosition + 3, 0).setText('L, ')

            self.comboBox_main_connect_type.setDisabled(True)

    def connect_type_choose(self):
        for i in range(0, self.tableWidget_2.rowCount(), 5):
            self.rowPosition = i
            self.tableWidget_2.cellWidget(self.rowPosition + 2, 3).setCurrentIndex(
                self.comboBox_main_connect_type.currentIndex())

    def sec_height_choose(self):
        rowPosition = self.tableWidget_2.currentRow()
        if self.tableWidget_2.cellWidget(rowPosition, 0).currentIndex() == 0:
            self.tableWidget_2.cellWidget(rowPosition, 2).removeItem(0)
            self.tableWidget_2.cellWidget(rowPosition, 2).addItem('160')
            self.tableWidget_2.cellWidget(rowPosition, 2).addItem('170')
            self.tableWidget_2.cellWidget(rowPosition, 2).addItem('180')
            self.tableWidget_2.cellWidget(rowPosition, 2).setCurrentIndex(2)
        if self.tableWidget_2.cellWidget(rowPosition, 0).currentIndex() == 1:
            self.tableWidget_2.cellWidget(rowPosition, 2).removeItem(2)
            self.tableWidget_2.cellWidget(rowPosition, 2).removeItem(1)
            self.tableWidget_2.cellWidget(rowPosition, 2).removeItem(0)
            self.tableWidget_2.cellWidget(rowPosition, 2).addItem('125')

    def main_type_pr_choose(self):
        for i in range(0, self.tableWidget_2.rowCount(), 5):
            self.rowPosition = i
            self.tableWidget_2.cellWidget(self.rowPosition, 4).setCurrentIndex(
                self.comboBox_main_pr_type.currentIndex())
            if self.comboBox_main_pr_type.currentIndex() == 2:
                self.comboBox_main_connect_type.setDisabled(True)
            else:
                self.comboBox_main_connect_type.setDisabled(False)
            self.type_pr_choose()

    def local_type_pr_choose(self):
        self.rowPosition = self.tableWidget_2.currentRow()
        self.type_pr_choose()

    def calc_pr_process(self):
        try:
            connect_type_massive = (0.97, 0.97, 0.78, 1, 0.8, 0.78)
            sec_dq_massive = (2.8, 2.7, 2.6, 2.25)
            panel_1_l_massive = (0.8977, -0.0466, 0.7424, 0.127, 0.5974, -0.0466, 0.5104, 0.169)
            panel_2_l_massive = (0.7168, -0.1399, 0.5637, 0.0862, 0.8298, -0.1503, 0.6366, -0.2692)
            panel_1_qn_massive = (1.114, 0.0466, 1.3469, -0.127, 1.674, 0.0466, 1.9593, -0.169)
            panel_2_qn_massive = (1.397, 0.1399, 1.7741, -0.0862, 2.2051, 0.1503, 2.5708, 0.2692)
            panel_1_dq_massive = (0.02, 0.0284, 0.0243, -0.0225, 0.0301, 0.0071, 0.0353, 0.0107)
            panel_2_dq_massive = (0.0251, 0.0238, 0.032, 0.0401, 0.0397, 0.0256, 0.0462, 0.0408)
            reg_pov_1_massive = (0.157, 0.22, 0.29, 0.372, 0.436, 0.529, 0.651, 0.779)
            reg_pov_2_massive = (0.131, 0.18, 0.238, 0.305, 0.357, 0.434, 0.558, 0.668)
            cp = 4200
            if self.checkBox_panel.isChecked():
                coef_panel = 1.2
            else:
                coef_panel = 1
            if self.checkBox_k_tr.isChecked() and self.comboBox_3_tube.currentText() == '':
                coef_t_tr = 1
            else:
                coef_t_tr = 0
            rowPosition = 2
            p = 0
            for i in range(self.tableWidget.rowCount()):
                if self.tableWidget.item(i, 3).text() == 't, C':
                    l_tr = self.ns_room_long[p] * 2
                    p += 1
                    t_vn = int(self.tableWidget.item(i, 4).text())
                    dt = float(round((self.t1 + self.t2) / 2 - t_vn))
                    q_t = float(round(coef_t_tr * (1.597 * dt - 27.571)))
                    Q_t = q_t * l_tr
                    Q_room = int(self.tableWidget.item(i, 5).text()) - Q_t
                    N_pr = int(self.tableWidget_2.item(rowPosition, 4).text())
                    Q_tr = int(round(Q_room / N_pr))
                    G_pred = float(round(3600 * Q_tr / (cp * (self.t1 - self.t2))))
                    Q_Qn_G_Gn = round(0.2695 * log(G_pred / 60) + 0.9912, 2)
                    Q_tr_g = round((Q_tr + (1 - Q_Qn_G_Gn) * Q_tr) * coef_panel)
                    if self.tableWidget_2.cellWidget(rowPosition - 2, 4).currentIndex() == 0:
                        Q_n_sec = int(self.tableWidget_2.cellWidget(rowPosition, 2).currentText())
                        if self.tableWidget_2.cellWidget(rowPosition, 0).currentIndex() == 0:
                            for m in range(3):
                                if self.tableWidget_2.cellWidget(rowPosition, 2).currentIndex() == m:
                                    dQ_sec = sec_dq_massive[m]
                        else:
                            dQ_sec = sec_dq_massive[3]
                        for m in range(6):
                            if self.tableWidget_2.cellWidget(rowPosition, 3).currentIndex() == m:
                                f = connect_type_massive[m]
                        Q_r_sec = round(f * (Q_n_sec - (70 - dt) * dQ_sec))
                        self.Gab = int(ceil(Q_tr_g / Q_r_sec))
                        self.Q_r = int(self.Gab * Q_r_sec)
                        self.Q_n = int(self.Gab * Q_n_sec)
                        self.G_pr = float(round(3600 * self.Q_r / (cp * (self.t1 - self.t2)), 1))
                        self.Q_r_room = int(self.Q_r * N_pr)
                    if self.tableWidget_2.cellWidget(rowPosition - 2, 4).currentIndex() == 1:
                        if self.tableWidget_2.cellWidget(rowPosition, 1).currentIndex() == 0:
                            for m in range(4):
                                if self.tableWidget_2.cellWidget(rowPosition, 0).currentIndex() == m:
                                    a_l = panel_1_l_massive[2 * m]
                                    b_l = panel_1_l_massive[2 * m + 1]
                                    a_qn = panel_1_qn_massive[2 * m]
                                    b_qn = panel_1_qn_massive[2 * m + 1]
                                    a_dq = panel_1_dq_massive[2 * m]
                                    b_dq = panel_1_dq_massive[2 * m + 1]
                        if self.tableWidget_2.cellWidget(rowPosition, 1).currentIndex() == 1:
                            for m in range(4):
                                if self.tableWidget_2.cellWidget(rowPosition, 0).currentIndex() == m:
                                    a_l = panel_2_l_massive[2 * m]
                                    b_l = panel_2_l_massive[2 * m + 1]
                                    a_qn = panel_2_qn_massive[2 * m]
                                    b_qn = panel_2_qn_massive[2 * m + 1]
                                    a_dq = panel_2_dq_massive[2 * m]
                                    b_dq = panel_2_dq_massive[2 * m + 1]
                        for m in range(6):
                            if self.tableWidget_2.cellWidget(rowPosition, 3).currentIndex() == m:
                                f = connect_type_massive[m]
                        Lpr_pan = round(a_l * Q_tr_g + b_l, -2)
                        Q_n_pan = round(a_qn * Lpr_pan + b_qn)
                        dQ_pan = round(a_dq * Lpr_pan + b_dq, 1)
                        Q_r_pan = round(f * (Q_n_pan - (70 - dt) * dQ_pan))
                        while Q_r_pan > Q_tr_g:
                            Lpr_pan = Lpr_pan - 100
                            Q_n_pan = round(a_qn * Lpr_pan + b_qn)
                            dQ_pan = round(a_dq * Lpr_pan + b_dq, 1)
                            Q_r_pan = round(f * (Q_n_pan - (70 - dt) * dQ_pan))
                        else:
                            Lpr_pan = Lpr_pan + 100
                            Q_n_pan = round(a_qn * Lpr_pan + b_qn)
                            dQ_pan = round(a_dq * Lpr_pan + b_dq, 1)
                            Q_r_pan = round(f * (Q_n_pan - (70 - dt) * dQ_pan))
                            self.tableWidget_2.item(rowPosition, 2).setText(str(Q_tr_g))
                            self.Gab = int(Lpr_pan)
                            self.Q_r = int(Q_r_pan)
                            self.Q_n = int(Q_n_pan)
                            self.G_pr = float(round(3600 * self.Q_r / (cp * (self.t1 - self.t2)), 1))
                            self.Q_r_room = int(self.Q_r * N_pr)
                    if self.tableWidget_2.cellWidget(rowPosition - 2, 4).currentIndex() == 2:
                        if self.tableWidget_2.cellWidget(rowPosition, 1).currentIndex() == 0:
                            for m in range(8):
                                if self.tableWidget_2.cellWidget(rowPosition, 0).currentIndex() == m:
                                    pov_reg = reg_pov_1_massive[m]
                        else:
                            for m in range(8):
                                if self.tableWidget_2.cellWidget(rowPosition, 0).currentIndex() == m:
                                    pov_reg = reg_pov_2_massive[m]

                        N_row = int(self.tableWidget_2.cellWidget(rowPosition, 1).currentText())
                        f_reg = round(0.86 * Q_tr / 435, 2)
                        self.tableWidget_2.item(rowPosition, 2).setText(str(Q_tr))
                        self.Gab = round(f_reg / pov_reg / N_row, 2)
                        self.Q_r = int(round(Q_tr / (dt / 64.5) ** 1.32))
                        self.Q_n = self.Q_r
                        self.G_pr = float(round(3600 * self.Q_r / (cp * (self.t1 - self.t2)), 1))
                        self.Q_r_room = int(self.Q_r * N_pr)

                    self.tableWidget_2.item(rowPosition + 2, 0).setText(str(self.Gab))
                    self.tableWidget_2.item(rowPosition + 2, 1).setText(str(self.Q_r))
                    self.tableWidget_2.item(rowPosition + 2, 2).setText(str(self.Q_n))
                    self.tableWidget_2.item(rowPosition + 2, 3).setText(str(self.G_pr))
                    self.tableWidget_2.item(rowPosition + 2, 4).setText(str(self.Q_r_room))
                    rowPosition += 5
        except ValueError:
            MyError.value_error_notify(self)
        except AttributeError:
            MyError.atrribute_error_notify(self)

    def sum_heat_pr(self):
        try:
            sum_heat_so_r = 0
            sum_heat_so_n = 0
            sum_g_so = 0
            for i in range(4, self.tableWidget_2.rowCount(), 5):
                sum_heat_so_r += int(self.tableWidget_2.item(i, 4).text())
                sum_heat_so_n = sum_heat_so_n + int(self.tableWidget_2.item(i, 2).text()) * int(
                    self.tableWidget_2.item(i - 2, 4).text())
                sum_g_so = sum_g_so + float(self.tableWidget_2.item(i, 3).text()) * int(
                    self.tableWidget_2.item(i - 2, 4).text())
                if i + 1 == self.tableWidget_2.rowCount():
                    self.lineEdit_18_sum_heat.setText(str(float(round(sum_heat_so_r / 1000, 2))))
                    self.lineEdit_22_sum_n_heat.setText(str(float(round(sum_heat_so_n / 1000, 2))))
                    self.lineEdit_20_sum_rate.setText(str(float(round(sum_g_so, 1))))
                    delta_heat = (1 - float(self.lineEdit_17_sum_h_loss.text()) / float(
                        round(sum_heat_so_r / 1000, 2))) * 100
                    delta_rate = (1 - float(self.lineEdit_24_n_rate.text()) / float(round(sum_g_so, 1))) * 100
                    self.lineEdit_23_dl_heat.setText(str(float(round(delta_heat, 2))))
                    self.lineEdit_25_dl_rate.setText(str(float(round(delta_rate, 2))))
        except ValueError:
            return
        except AttributeError:
            return

    def delete_pr_table(self):
        self.pushButton_add_row.setDisabled(False)
        self.pushButton_3_del_row.setDisabled(False)
        self.pushButton_2_room.setDisabled(False)
        self.pushButton_4_calc_h_loss.setDisabled(False)
        self.comboBox_main_pr_type.setDisabled(True)
        for i in range(self.tableWidget_2.rowCount()):
            rowPosition = self.tableWidget_2.rowCount()
            self.tableWidget_2.removeRow(rowPosition - 1)

    def add_section(self):
        if self.lineEdit_add_sec.text() == '':
            colPosition = self.tableWidget_3.columnCount()
        else:
            try:
                colPosition = int(self.lineEdit_add_sec.text())
                self.lineEdit_add_sec.clear()
            except ValueError:
                MyError.value_error_notify(self)
        self.tableWidget_3.insertColumn(colPosition)
        for i in range(5):
            self.sec_item = QtWidgets.QTableWidgetItem('')
            self.sec_item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.tableWidget_3.setItem(i, colPosition, self.sec_item)
        for i in range(5, 10):
            self.sec_item = QtWidgets.QTableWidgetItem('0')
            self.sec_item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.tableWidget_3.setItem(i, colPosition, self.sec_item)
        self.sec_item = QtWidgets.QTableWidgetItem('0')
        self.sec_item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
        self.sec_item.setBackground(QtGui.QColor(0, 170, 255))
        self.tableWidget_3.setItem(10, colPosition, self.sec_item)
        for i in range(11, 18):
            self.sec_item = QtWidgets.QTableWidgetItem('')
            self.sec_item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.tableWidget_3.setItem(i, colPosition, self.sec_item)
            self.sec_item.setFlags(QtCore.Qt.ItemIsEnabled)
        self.tableWidget_3.resizeColumnsToContents()

    def del_section(self):
        try:
            colPosition = self.tableWidget_3.columnCount() - 1
            if self.lineEdit_f_d_sec.text() == "" or self.lineEdit_l_d_sec.text() == "":
                self.tableWidget_3.removeColumn(colPosition)
            else:
                f_int_col = int(self.lineEdit_f_d_sec.text()) - 1
                l_int_col = int(self.lineEdit_l_d_sec.text())
                for i in range(f_int_col, l_int_col):
                    self.tableWidget_3.removeColumn(f_int_col)
                self.lineEdit_f_d_sec.clear()
                self.lineEdit_l_d_sec.clear()
        except ValueError:
            MyError.value_error_notify(self)

    #     
    def nearest_value(self, x, row):
        t = x
        for i in range(len(row)):
            delta = abs(x - row[i])
            if delta < t:
                t = delta
                value = i
        return value

    def calc_hydro_process(self):
        try:
            #   
            hs = HSDiag()
            visc = Visc()
            t_Av = (self.t1 + self.t2) / 2
            if self.comboBox_4_p_n.currentIndex() == 1:
                p_In = round(float(self.lineEdit_p_n.text()) * 9.806, 1)
            elif self.comboBox_4_p_n.currentIndex() == 2:
                p_In = round(float(self.lineEdit_p_n.text()) * 101.325, 1)
            else:
                p_In = float(self.lineEdit_p_n.text())
            PhysChar_Av = hs.props_tp(t_Av, p_In * 1000)
            self.ro_Av = 1 / PhysChar_Av['v']
            KinVis_Av = visc.kvisc_tp(t_Av, p_In * 1000)

            otv_massive_st = (64, 0.3, 40, 0.4, 32, 0.5, 25, 0.5, 20, 0.6, 15, 0.8)
            otv_massive_pp = (64, 1.1, 40, 1.6, 32, 1.8, 25, 2, 20, 2.1, 15, 2.2)
            if self.comboBox_3_tube.currentIndex() == 0:
                otv_massive = otv_massive_st
                tr_p = 0.5
                tr_s_r = 4.5
                k_ecv = 0.3
                prh = 0.1
            if self.comboBox_3_tube.currentIndex() == 1:
                otv_massive = otv_massive_pp
                tr_p = 1
                tr_s_r = 2
                k_ecv = 0.01
                prh = 0.4
            kshz = 0.1

            for i in range(self.tableWidget_3.columnCount()):
                L_sec = float(self.tableWidget_3.item(1, i).text())
                D_in = float(self.tableWidget_3.item(2, i).text())
                G_sec = float(self.tableWidget_3.item(3, i).text())
                N_otv = int(self.tableWidget_3.item(5, i).text())
                N_tr_p = int(self.tableWidget_3.item(6, i).text())
                N_tr_s_r = int(self.tableWidget_3.item(7, i).text())
                N_prh = int(self.tableWidget_3.item(8, i).text())
                N_kshz = int(self.tableWidget_3.item(9, i).text())
                Other_ms = float(self.tableWidget_3.item(10, i).text())
                otv_ms = otv_massive[2 * self.nearest_value(D_in, otv_massive[::2]) + 1]
                dzeta = round(
                    N_otv * otv_ms + N_tr_p * tr_p + N_tr_s_r * 0.5 * tr_s_r + N_kshz * kshz + N_prh * prh + Other_ms,
                    2)
                A = 8 / (3600 ** 2 * pi ** 2 * (D_in / 1000) ** 4 * self.ro_Av)
                w = round(4 * G_sec / (3600 * self.ro_Av * pi * (D_in / 1000) ** 2), 2)
                Re = round(w * D_in / 1000 / KinVis_Av)
                lamda = round(0.11 * (k_ecv / D_in + 68 / Re) ** 0.25, 4)
                lamda_ms = round(lamda * L_sec / (D_in / 1000), 2)
                S = round((dzeta + lamda_ms) * A, 6)
                dp = round(2 * (S * G_sec ** 2), 1)
                self.tableWidget_3.item(11, i).setText(str(w))
                self.tableWidget_3.setAlternatingRowColors(True)
                if w > 0.6:
                    self.tableWidget_3.item(2, i).setBackground(QtGui.QColor(255, 90, 0))
                    self.tableWidget_3.item(11, i).setBackground(QtGui.QColor(255, 90, 0))
                else:
                    if self.comboBox_wind_style.currentIndex() == -1:
                        self.tableWidget_3.item(2, i).setBackground(QtGui.QColor(255, 255, 255))
                        self.tableWidget_3.item(11, i).setBackground(QtGui.QColor(240, 240, 240))
                    if self.comboBox_wind_style.currentIndex() == 0:
                        self.tableWidget_3.item(2, i).setBackground(QtGui.QColor(200, 200, 200))
                        self.tableWidget_3.item(11, i).setBackground(QtGui.QColor(200, 200, 200))
                    if self.comboBox_wind_style.currentIndex() == 1:
                        self.tableWidget_3.item(2, i).setBackground(QtGui.QColor(215, 215, 215))
                        self.tableWidget_3.item(11, i).setBackground(QtGui.QColor(215, 215, 215))
                    if self.comboBox_wind_style.currentIndex() == 2:
                        self.tableWidget_3.item(2, i).setBackground(QtGui.QColor(230, 230, 230))
                        self.tableWidget_3.item(11, i).setBackground(QtGui.QColor(230, 230, 230))
                self.tableWidget_3.item(12, i).setText(str(Re))
                self.tableWidget_3.item(13, i).setText(str(dzeta))
                self.tableWidget_3.item(14, i).setText(str(lamda))
                self.tableWidget_3.item(15, i).setText(str(S))
                self.tableWidget_3.item(16, i).setText(str(dp))
                self.tableWidget_3.resizeColumnsToContents()
            a = False
            while True:
                if a == False:
                    for i in range(self.tableWidget_3.columnCount()):
                        if self.tableWidget_3.item(4, i).text() == '' or self.tableWidget_3.item(4, i).text() == '-':
                            self.tableWidget_3.item(4, i).setText('')
                            sum_dp = self.tableWidget_3.item(16, i).text()
                        for m in range(self.tableWidget_3.columnCount()):
                            if self.tableWidget_3.item(0, m).text() == self.tableWidget_3.item(4, i).text():
                                try:
                                    sum_dp = round(float(self.tableWidget_3.item(16, i).text()) +
                                                   float(self.tableWidget_3.item(17, m).text()), 1)
                                except:
                                    sum_dp = ''
                                    continue
                        self.tableWidget_3.item(17, i).setText(str(sum_dp))
                    a = True
                    for i in range(self.tableWidget_3.columnCount()):
                        if self.tableWidget_3.item(17, i).text() == '':
                            a = False
                else:
                    break
        except ValueError:
            MyError.value_error_notify(self)
        except AttributeError:
            MyError.value_error_notify(self)

    def sum_pressure_loss(self):
        try:
            sum_dp_list = []
            not_vck = []
            self.vck = []
            for i in range(self.tableWidget_3.columnCount()):
                self.tableWidget_3.item(17, i).setBackground(QtGui.QColor(240, 240, 240))
                sum_dp_list.append(float(self.tableWidget_3.item(17, i).text()))
            dp_gck = round(max(sum_dp_list) / 1000, 1)
            self.lineEdit_21_sum_p_loss.setText(str(dp_gck))
            a = max(range(len(sum_dp_list)), key=lambda i: sum_dp_list[i])
            self.gck_ind = a
            for i in range(self.tableWidget_3.columnCount()):
                for m in range(self.tableWidget_3.columnCount()):
                    if self.tableWidget_3.item(0, m).text() == self.tableWidget_3.item(4, a).text():
                        a = m
            self.lineEdit_19_name_gck.setText(
                self.tableWidget_3.item(0, a).text() + '  ' + self.tableWidget_3.item(0, self.gck_ind).text())
            self.tableWidget_3.item(17, self.gck_ind).setBackground(QtGui.QColor(0, 210, 0))
            for i in range(self.tableWidget_3.columnCount()):
                for m in range(self.tableWidget_3.columnCount()):
                    if self.tableWidget_3.item(0, i).text() == self.tableWidget_3.item(4, m).text():
                        not_vck.append(i)
            not_vck = list(set(not_vck))
            vck_quan = self.tableWidget_3.columnCount() - len(not_vck)
            for el in range(len(not_vck)):
                try:
                    if not_vck[el] + 1 != not_vck[el + 1]:
                        n = not_vck[el + 1] - not_vck[el]
                        for m in range(1, n):
                            k = not_vck[el] + m
                            self.vck.append(k)
                            vck_quan = vck_quan - 1
                except:
                    c = vck_quan
                    for m in range(0, c):
                        k = not_vck[el] + m + 1
                        self.vck.append(k)
                        vck_quan = vck_quan - 1
            for i in range(len(self.vck)):
                if self.vck[i] == self.gck_ind:
                    self.vck.pop(i)
                    break
            for i in range(len(self.vck)):
                self.tableWidget_3.item(17, self.vck[i]).setBackground(QtGui.QColor(0, 255, 127))
            self.pushButton_form_balance.setDisabled(False)
        except ValueError:
            return

    def get_data_from_hydro(self):
        for i in range(len(self.vck)):
            self.tableWidget_4.insertRow(i)
            for m in range(self.tableWidget_4.columnCount()):
                self.bal_item = QtWidgets.QTableWidgetItem('')
                self.bal_item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.bal_item.setFlags(QtCore.Qt.ItemIsEnabled)
                self.tableWidget_4.setItem(i, m, self.bal_item)
            self.tableWidget_4.item(i, 0).setText(str(self.vck[i] + 1))
            self.tableWidget_4.item(i, 1).setText(self.tableWidget_3.item(0, self.vck[i]).text())
            self.tableWidget_4.resizeColumnsToContents()
        self.pushButton_add_section.setDisabled(True)
        self.pushButton_del_section.setDisabled(True)
        self.pushButton_hydro_calc.setDisabled(True)
        self.pushButton_form_balance.setDisabled(True)
        self.pushButton_calc_balance.setDisabled(False)

    def delete_balance_table(self):
        self.pushButton_add_section.setDisabled(False)
        self.pushButton_del_section.setDisabled(False)
        self.pushButton_hydro_calc.setDisabled(False)
        self.pushButton_calc_balance.setDisabled(True)
        for i in range(self.tableWidget_4.rowCount()):
            rowPosition = self.tableWidget_4.rowCount()
            self.tableWidget_4.removeRow(rowPosition - 1)

    def calc_balance_process(self):
        try:
            dp_gck = float(self.tableWidget_3.item(17, self.gck_ind).text())
            for i in range(len(self.vck)):
                dp_bal = dp_gck - float(self.tableWidget_3.item(17, self.vck[i]).text())
                d_dr_sh = round(3.16 * float(self.tableWidget_3.item(3, self.vck[i]).text()) ** 0.5 / dp_bal ** 0.25)
                kvs = round(
                    float(self.tableWidget_3.item(3, self.vck[i]).text()) / self.ro_Av / (dp_bal * 10 ** -5) ** 0.5, 1)
                self.tableWidget_4.item(i, 2).setText(str(d_dr_sh))
                self.tableWidget_4.item(i, 3).setText(str(kvs))
                if d_dr_sh <= 3:
                    self.tableWidget_4.item(i, 2).setBackground(QtGui.QColor(255, 90, 0))
                if kvs < 1:
                    self.tableWidget_4.item(i, 3).setBackground(QtGui.QColor(255, 90, 0))
        except ValueError:
            MyError.value_error_notify(self)

    def tube_add_column(self):
        if self.lineEdit_add_col.text() == '':
            colPosition = self.tableWidget_5.columnCount()
        else:
            try:
                colPosition = int(self.lineEdit_add_col.text())
                self.lineEdit_add_col.clear()
            except ValueError:
                MyError.value_error_notify(self)
        self.tableWidget_5.insertColumn(colPosition)

        self.tube_purpose = QtWidgets.QComboBox()
        self.tube_purpose.setStyleSheet("background-color: rgb(170, 255, 127);")
        self.tableWidget_5.setCellWidget(0, colPosition, self.tube_purpose)
        self.tableWidget_5.cellWidget(0, colPosition).addItem('  ')
        self.tableWidget_5.cellWidget(0, colPosition).addItem('')
        self.tableWidget_5.cellWidget(0, colPosition).addItem('')
        self.tube_purpose.currentIndexChanged.connect(self.tube_hvs_mod)

        self.tube_type = QtWidgets.QComboBox()
        self.tube_type.setStyleSheet("background-color: rgb(170, 255, 127);")
        self.tableWidget_5.setCellWidget(1, colPosition, self.tube_type)
        self.tableWidget_5.cellWidget(1, colPosition).addItem(' 10704-91')
        self.tableWidget_5.cellWidget(1, colPosition).addItem(' 3262-75')
        self.tableWidget_5.cellWidget(1, colPosition).addItem(' 8732-78')
        if self.comboBox_q_r_choose.currentIndex() == 0:
            self.col_item = QtWidgets.QTableWidgetItem()
            self.col_item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.col_item.setBackground(QtGui.QColor(85, 255, 255))
            self.tableWidget_5.setItem(2, colPosition, self.col_item)
            self.col_item = QtWidgets.QTableWidgetItem()
            self.col_item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.col_item.setFlags(QtCore.Qt.ItemIsEnabled)
            self.tableWidget_5.setItem(3, colPosition, self.col_item)
        elif self.comboBox_q_r_choose.currentIndex() == 1:
            self.col_item = QtWidgets.QTableWidgetItem()
            self.col_item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.col_item.setBackground(QtGui.QColor(85, 255, 255))
            self.tableWidget_5.setItem(3, colPosition, self.col_item)
            self.col_item = QtWidgets.QTableWidgetItem()
            self.col_item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.col_item.setFlags(QtCore.Qt.ItemIsEnabled)
            self.tableWidget_5.setItem(2, colPosition, self.col_item)
        for i in range(4, 7):
            self.col_item = QtWidgets.QTableWidgetItem()
            self.col_item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.col_item.setFlags(QtCore.Qt.ItemIsEnabled)
            self.tableWidget_5.setItem(i, colPosition, self.col_item)
        self.col_item = QtWidgets.QTableWidgetItem()
        self.col_item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
        self.col_item.setBackground(QtGui.QColor(85, 255, 255))
        self.tableWidget_5.setItem(7, colPosition, self.col_item)
        for i in range(8, 10):
            self.col_item = QtWidgets.QTableWidgetItem()
            self.col_item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.col_item.setFlags(QtCore.Qt.ItemIsEnabled)
            self.tableWidget_5.setItem(i, colPosition, self.col_item)
        for i in range(10, 12):
            self.col_item = QtWidgets.QTableWidgetItem()
            self.col_item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.col_item.setBackground(QtGui.QColor(85, 255, 255))
            self.tableWidget_5.setItem(i, colPosition, self.col_item)
        for i in range(12, 17):
            self.col_item = QtWidgets.QTableWidgetItem()
            self.col_item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.col_item.setFlags(QtCore.Qt.ItemIsEnabled)
            self.tableWidget_5.setItem(i, colPosition, self.col_item)
        self.tableWidget_5.resizeColumnsToContents()

    def tube_del_column(self):
        try:
            colPosition = self.tableWidget_5.columnCount() - 1
            if self.lineEdit_f_d_col.text() == "" or self.lineEdit_l_d_col.text() == "":
                self.tableWidget_5.removeColumn(colPosition)
            else:
                f_int_col = int(self.lineEdit_f_d_col.text()) - 1
                l_int_col = int(self.lineEdit_l_d_col.text())
                for i in range(f_int_col, l_int_col):
                    self.tableWidget_5.removeColumn(f_int_col)
                self.lineEdit_f_d_col.clear()
                self.lineEdit_l_d_col.clear()
        except ValueError:
            MyError.value_error_notify(self)

    def tube_main_calc_way(self):
        for i in range(self.tableWidget_5.columnCount()):
            self.colPosition = i
            self.tube_calc_way()

    def tube_calc_way(self):
        if self.tableWidget_5.cellWidget(0, self.colPosition).currentIndex() != 0:
            pass
        elif self.comboBox_q_r_choose.currentIndex() == 0:
            self.col_item = QtWidgets.QTableWidgetItem()
            self.col_item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.col_item.setBackground(QtGui.QColor(85, 255, 255))
            self.tableWidget_5.setItem(2, self.colPosition, self.col_item)
            self.col_item = QtWidgets.QTableWidgetItem()
            self.col_item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.col_item.setFlags(QtCore.Qt.ItemIsEnabled)
            if self.comboBox_wind_style.currentIndex() == -1:
                self.col_item.setBackground(QtGui.QColor(240, 240, 240))
            if self.comboBox_wind_style.currentIndex() == 0:
                self.col_item.setBackground(QtGui.QColor(200, 200, 200))
            if self.comboBox_wind_style.currentIndex() == 1:
                self.col_item.setBackground(QtGui.QColor(215, 215, 215))
            if self.comboBox_wind_style.currentIndex() == 2:
                self.col_item.setBackground(QtGui.QColor(230, 230, 230))
            self.tableWidget_5.setItem(3, self.colPosition, self.col_item)
        elif self.comboBox_q_r_choose.currentIndex() == 1:
            self.col_item = QtWidgets.QTableWidgetItem()
            self.col_item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.col_item.setBackground(QtGui.QColor(85, 255, 255))
            self.tableWidget_5.setItem(3, self.colPosition, self.col_item)
            self.col_item = QtWidgets.QTableWidgetItem()
            self.col_item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.col_item.setFlags(QtCore.Qt.ItemIsEnabled)
            if self.comboBox_wind_style.currentIndex() == -1:
                self.col_item.setBackground(QtGui.QColor(240, 240, 240))
            if self.comboBox_wind_style.currentIndex() == 0:
                self.col_item.setBackground(QtGui.QColor(200, 200, 200))
            if self.comboBox_wind_style.currentIndex() == 1:
                self.col_item.setBackground(QtGui.QColor(215, 215, 215))
            if self.comboBox_wind_style.currentIndex() == 2:
                self.col_item.setBackground(QtGui.QColor(230, 230, 230))
            self.tableWidget_5.setItem(2, self.colPosition, self.col_item)

    def tube_hvs_mod(self):
        colPosition = self.tableWidget_5.currentColumn()
        if self.tableWidget_5.cellWidget(0, colPosition).currentIndex() != 0:
            self.col_item = QtWidgets.QTableWidgetItem()
            self.col_item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.col_item.setFlags(QtCore.Qt.ItemIsEnabled)
            self.tableWidget_5.setItem(2, colPosition, self.col_item)
            self.col_item = QtWidgets.QTableWidgetItem()
            self.col_item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.col_item.setFlags(QtCore.Qt.ItemIsEnabled)
            self.tableWidget_5.setItem(3, colPosition, self.col_item)
            self.col_item = QtWidgets.QTableWidgetItem()
            self.col_item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.col_item.setBackground(QtGui.QColor(85, 255, 255))
            self.tableWidget_5.setItem(5, colPosition, self.col_item)
        else:
            self.col_item = QtWidgets.QTableWidgetItem()
            self.col_item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.col_item.setFlags(QtCore.Qt.ItemIsEnabled)
            self.tableWidget_5.setItem(5, colPosition, self.col_item)
            self.colPosition = colPosition
            self.tube_calc_way()

    def tube_calc_process(self):
        try:
            hs = HSDiag()
            DN_massive = [15, 20, 25, 32, 40, 50, 65, 80, 100, 125, 150, 200, 250, 300]
            dn_10704_massive = [20, 25, 32, 38, 45, 57, 76, 89, 108, 133, 159, 219, 273, 325]
            dM_10704_massive = [0.808, 1.13, 1.48, 2.62, 4, 5.4, 7.38, 10.26, 12.73, 17.15, 26.39, 39.51, 54.9]
            dn_3262_massive = [21.3, 26.8, 33.5, 42.3, 48, 60, 75.5, 88.5, 114, 140, 165]
            dM_3262_massive = [1.28, 1.66, 2.39, 3.09, 3.84, 4.88, 7.05, 8.34, 12.15, 15.04, 17.81]
            dn_8732_massive = ['', '', '', '', 45, 57, 76, 89, 108, 133, 159, 219, 273, 325]
            dM_8732_massive = ['', '', '', '', 3.58, 4.62, 6.26, 8.39, 10.26, 14.26, 18.99, 31.52, 45.92, 62.54]
            S_n_sum = 0
            S_iz_sum = 0
            v_sum = 0
            M_sum = 0
            for i in range(self.tableWidget_5.columnCount()):
                if self.tableWidget_5.cellWidget(0, i).currentIndex() == 0:
                    if self.comboBox_q_r_choose.currentIndex() == 0:
                        Q_r = float(self.tableWidget_5.item(2, i).text())
                        self.tableWidget_5.item(3, i).setText(str(round(Q_r * 860)))
                    elif self.comboBox_q_r_choose.currentIndex() == 1:
                        Q_r = float(self.tableWidget_5.item(3, i).text()) * 0.001163
                        self.tableWidget_5.item(2, i).setText(str(round(Q_r, 2)))
                w = float(self.tableWidget_5.item(7, i).text())
                if self.tableWidget_5.cellWidget(0, i).currentIndex() == 0 and w > 0.6:
                    self.tableWidget_5.item(7, i).setBackground(QtGui.QColor(255, 90, 0))
                    self.tableWidget_5.item(8, i).setBackground(QtGui.QColor(255, 90, 0))
                    self.tableWidget_5.item(9, i).setBackground(QtGui.QColor(255, 90, 0))
                elif w > 1.2:
                    self.tableWidget_5.item(7, i).setBackground(QtGui.QColor(255, 90, 0))
                    self.tableWidget_5.item(8, i).setBackground(QtGui.QColor(255, 90, 0))
                    self.tableWidget_5.item(9, i).setBackground(QtGui.QColor(255, 90, 0))
                else:
                    self.tableWidget_5.item(7, i).setBackground(QtGui.QColor(85, 255, 255))
                    if self.comboBox_wind_style.currentIndex() == -1:
                        self.tableWidget_5.item(8, i).setBackground(QtGui.QColor(255, 255, 255))
                        self.tableWidget_5.item(9, i).setBackground(QtGui.QColor(240, 240, 240))
                    if self.comboBox_wind_style.currentIndex() == 0:
                        self.tableWidget_5.item(8, i).setBackground(QtGui.QColor(200, 200, 200))
                        self.tableWidget_5.item(9, i).setBackground(QtGui.QColor(200, 200, 200))
                    if self.comboBox_wind_style.currentIndex() == 1:
                        self.tableWidget_5.item(8, i).setBackground(QtGui.QColor(215, 215, 215))
                        self.tableWidget_5.item(9, i).setBackground(QtGui.QColor(215, 215, 215))
                    if self.comboBox_wind_style.currentIndex() == 2:
                        self.tableWidget_5.item(8, i).setBackground(QtGui.QColor(230, 230, 230))
                        self.tableWidget_5.item(9, i).setBackground(QtGui.QColor(230, 230, 230))
                if self.tableWidget_5.cellWidget(0, i).currentIndex() == 0:
                    delta_t = self.t1 - self.t2
                    t_Av = (self.t1 + self.t2) / 2
                    if self.lineEdit_p_n.text() == '':
                        p_In = 400
                    else:
                        if self.comboBox_4_p_n.currentIndex() == 1:
                            p_In = round(float(self.lineEdit_p_n.text()) * 9.806, 1)
                        elif self.comboBox_4_p_n.currentIndex() == 2:
                            p_In = round(float(self.lineEdit_p_n.text()) * 101.325, 1)
                        elif self.comboBox_4_p_n.currentIndex() == 0:
                            p_In = float(self.lineEdit_p_n.text())
                elif self.tableWidget_5.cellWidget(0, i).currentIndex() == 1:
                    delta_t = 60
                    t_Av = 65
                    p_In = 400
                else:
                    delta_t = 15
                    t_Av = delta_t
                    p_In = 400
                PhysChar_Av = hs.props_tp(t_Av, p_In * 1000)
                ro_Av = 1 / PhysChar_Av['v']
                if self.tableWidget_5.cellWidget(0, i).currentIndex() == 0:
                    G_r = round(3600 * Q_r / (4.19 * delta_t))
                else:
                    V_s_cell = float(self.tableWidget_5.item(5, i).text())
                    G_r = round(V_s_cell * 3.6 * ro_Av)
                if self.tableWidget_5.cellWidget(0, i).currentIndex() == 1:
                    Q_r = round(G_r / 3600 * 4.19 * delta_t, 2)
                    self.tableWidget_5.item(2, i).setText(str(Q_r))
                    self.tableWidget_5.item(3, i).setText(str(round(Q_r * 860)))
                V_r = round(G_r / ro_Av, 2)
                V_s = G_r / (3600 * ro_Av)
                V_s_cell = round(V_s * 1000, 3)
                self.tableWidget_5.item(6, i).setText(str(V_r))
                self.tableWidget_5.item(4, i).setText(str(G_r))
                self.tableWidget_5.item(5, i).setText(str(V_s_cell))
                d_r = round(2 * sqrt(V_s / (pi * w)), 3)
                self.tableWidget_5.item(8, i).setText(str(d_r))
                DN = DN_massive[self.nearest_value(d_r * 1000, DN_massive)]
                self.tableWidget_5.item(9, i).setText(str(DN))
                if self.tableWidget_5.item(10, i).text() != '':
                    l = float(self.tableWidget_5.item(10, i).text())
                    if self.tableWidget_5.cellWidget(1, i).currentIndex() == 0:
                        D_n = dn_10704_massive[self.nearest_value(d_r * 1000, DN_massive)]
                        m = dM_10704_massive[self.nearest_value(d_r * 1000, DN_massive)]
                        self.tableWidget_5.cellWidget(1, i).setStyleSheet("background-color: rgb(170, 255, 127);")
                    elif self.tableWidget_5.cellWidget(1, i).currentIndex() == 1:
                        if DN > 150:
                            self.tableWidget_5.cellWidget(1, i).setStyleSheet("background-color: rgb(255, 90, 0);")
                            self.tableWidget_5.item(8, i).setBackground(QtGui.QColor(255, 90, 0))
                            self.tableWidget_5.item(9, i).setBackground(QtGui.QColor(255, 90, 0))
                            DN = DN_massive[10]
                            D_n = dn_3262_massive[10]
                            m = dM_3262_massive[10]
                        else:
                            D_n = dn_3262_massive[self.nearest_value(d_r * 1000, DN_massive)]
                            m = dM_3262_massive[self.nearest_value(d_r * 1000, DN_massive)]
                            self.tableWidget_5.cellWidget(1, i).setStyleSheet("background-color: rgb(170, 255, 127);")
                    elif self.tableWidget_5.cellWidget(1, i).currentIndex() == 2:
                        if DN < 40:
                            self.tableWidget_5.cellWidget(1, i).setStyleSheet("background-color: rgb(255, 90, 0);")
                            self.tableWidget_5.item(8, i).setBackground(QtGui.QColor(255, 90, 0))
                            self.tableWidget_5.item(9, i).setBackground(QtGui.QColor(255, 90, 0))
                            DN = DN_massive[4]
                            D_n = dn_8732_massive[4]
                            m = dM_8732_massive[4]
                        else:
                            D_n = dn_8732_massive[self.nearest_value(d_r * 1000, DN_massive)]
                            m = dM_8732_massive[self.nearest_value(d_r * 1000, DN_massive)]
                            self.tableWidget_5.cellWidget(1, i).setStyleSheet("background-color: rgb(170, 255, 127);")
                    self.tableWidget_5.item(9, i).setText(str(DN))
                    S_n = round(pi * D_n / 1000 * l, 3)
                    self.tableWidget_5.item(12, i).setText(str(S_n))
                    S_n_sum += S_n
                    v = round(pi / 4 * l * (DN / 1000) ** 2, 3)
                    self.tableWidget_5.item(14, i).setText(str(v))
                    v_sum += v
                    M = round(m * l + ro_Av * v, 2)
                    self.tableWidget_5.item(15, i).setText(str(M))
                    M_sum += M
                    M_l = round(m + ro_Av * v / l, 2)
                    self.tableWidget_5.item(16, i).setText(str(M_l))
                    if self.tableWidget_5.item(11, i).text() != '':
                        s_iz = float(self.tableWidget_5.item(11, i).text())
                        S_iz = round(pi * (D_n + s_iz) / 1000 * l, 2)
                        S_iz_sum += S_iz
                        self.tableWidget_5.item(13, i).setText(str(S_iz))
                else:
                    continue
            self.lineEdit_s_n_sum.setText(str(round(S_n_sum, 2)))
            self.lineEdit_s_iz_sum.setText(str(round(S_iz_sum, 1)))
            self.lineEdit_v_sum.setText(str(round(v_sum, 2)))
            self.lineEdit_m_sum.setText(str(round(M_sum, 1)))
        except ValueError:
            MyError.value_error_notify(self)
        except AttributeError:
            MyError.value_error_notify(self)

    def save_main_data(self):
        self.main_data = (self.lineEdit_5_obj_name.text(), self.lineEdit_6_t_n.text(), self.lineEdit_11_pol_1.text(),
                          self.lineEdit_13_pol_2.text(), self.lineEdit_7_pol_3.text(), self.lineEdit_14_pol_4.text(),
                          self.lineEdit_9_ns_16.text(), self.lineEdit_9_ns_20.text(), self.lineEdit_15_kr.text(),
                          self.lineEdit_12_wind.text(), self.lineEdit_16_door.text(), self.lineEdit_gate.text(),
                          self.checkBox_3_n_coef.isChecked(), self.lineEdit_p_n.text(), self.comboBox_t1.currentIndex(),
                          self.comboBox_2_t2.currentIndex(), self.comboBox_4_p_n.currentIndex(),
                          self.comboBox_3_tube.currentIndex(), self.lineEdit_17_sum_h_loss.text(),
                          self.lineEdit_18_sum_heat.text(), self.lineEdit_20_sum_rate.text(),
                          self.lineEdit_19_name_gck.text(), self.lineEdit_21_sum_p_loss.text(),
                          self.lineEdit_22_sum_n_heat.text(), self.lineEdit_23_dl_heat.text(),
                          self.lineEdit_24_n_rate.text(), self.lineEdit_25_dl_rate.text(),
                          self.lineEdit_another_ok.text(), self.ns_room_long, self.lineEdit_s_n_sum.text(),
                          self.lineEdit_s_iz_sum.text(), self.lineEdit_v_sum.text(), self.lineEdit_m_sum.text())

    def save_heat_data(self):
        self.row_heat_text = []
        for i in range(self.tableWidget.rowCount()):
            for m in range(self.tableWidget.columnCount()):
                current_item = self.tableWidget.item(i, m)
                if isinstance(current_item, QtWidgets.QTableWidgetItem):
                    current_text = self.tableWidget.item(i, m).text()
                else:
                    current_text = self.tableWidget.cellWidget(i, m).currentIndex()
                self.row_heat_text.append(current_text)

    def save_pr_data(self):
        self.row_pr_text = []
        current_main_pr_type = self.comboBox_main_pr_type.currentIndex()
        for i in range(self.tableWidget_2.rowCount()):
            for m in range(self.tableWidget_2.columnCount()):
                current_item = self.tableWidget_2.item(i, m)
                if isinstance(current_item, QtWidgets.QTableWidgetItem):
                    current_text = self.tableWidget_2.item(i, m).text()
                else:
                    current_text = self.tableWidget_2.cellWidget(i, m).currentIndex()
                self.row_pr_text.append(current_text)
        self.row_pr_text.append(current_main_pr_type)

    def save_hydro_data(self):
        self.row_hydro_text = []
        for i in range(self.tableWidget_3.columnCount()):
            for m in range(self.tableWidget_3.rowCount()):
                current_text = self.tableWidget_3.item(m, i).text()
                self.row_hydro_text.append(current_text)

    def save_balance_data(self):
        self.row_balance_text = []
        for i in range(self.tableWidget_4.columnCount()):
            for m in range(self.tableWidget_4.rowCount()):
                current_text = self.tableWidget_4.item(m, i).text()
                self.row_balance_text.append(current_text)
        try:
            self.row_balance_text.append(self.vck)
        except:
            pass

    def save_tube_data(self):
        self.row_tube_text = []
        for i in range(self.tableWidget_5.columnCount()):
            for m in range(self.tableWidget_5.rowCount()):
                current_item = self.tableWidget_5.item(m, i)
                if isinstance(current_item, QtWidgets.QTableWidgetItem):
                    current_text = self.tableWidget_5.item(m, i).text()
                else:
                    current_text = self.tableWidget_5.cellWidget(m, i).currentIndex()
                self.row_tube_text.append(current_text)

    def json_save_data(self):
        MyError.error = QMessageBox()
        MyError.error.setWindowTitle(" ")
        MyError.error.setIcon(QMessageBox.Question)
        MyError.error.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        path = os.path.join("data", "icon.png")
        icon = QtGui.QIcon(path)
        MyError.error.setWindowIcon(icon)
        json_data = {"main": self.main_data, "heat": self.row_heat_text, "pr": self.row_pr_text,
                     "hydro": self.row_hydro_text, "balance": self.row_balance_text, "tube": self.row_tube_text, "file_list": self.file_list}
        if self.lineEdit_save_as.text() == '':
            MyError.error.setText(' ?')
            save_confirm = MyError.error.exec_()
            if save_confirm == QMessageBox.Ok:
                path = os.path.join("data", "SVO_data.json")
                with open(path, "w") as write_file:
                    json.dump(json_data, write_file, indent=4)
                if os.path.isfile(path):
                    MyError.save_success(self)
                else:
                    MyError.save_fail(self)
            else:
                return
        else:
            file_name = self.lineEdit_5_obj_name.text() + '_' + self.lineEdit_save_as.text() + '.json'
            file_name_for_list = self.lineEdit_5_obj_name.text() + '_' + self.lineEdit_save_as.text()
            path = os.path.join("data", file_name)
            path_1 = os.path.join("data", "SVO_data.json")
            MyError.error.setText('    \{0}\?'.format(file_name_for_list))
            save_confirm = MyError.error.exec_()
            if save_confirm == QMessageBox.Ok:
                if os.path.isfile(path):
                    MyError.error.setText('     .   ?')
                    save_confirm = MyError.error.exec_()
                    if save_confirm == QMessageBox.Ok:
                        for i in range(0, self.comboBox_load_from.count()):
                            if file_name_for_list == self.comboBox_load_from.itemText(i):
                                self.comboBox_load_from.removeItem(i)
                        for i in range(0, len(self.file_list)):
                            if file_name_for_list == self.file_list[i]:
                                self.file_list.pop(i)
                                break
                        os.remove(path)
                        with open(path, "w") as write_file:
                            json.dump(json_data, write_file, indent=4)
                        if os.path.isfile(path):
                            self.comboBox_load_from.addItem(file_name_for_list)
                            self.file_list.append(file_name_for_list)
                            json_data = {"main": self.main_data, "heat": self.row_heat_text, "pr": self.row_pr_text,
                                         "hydro": self.row_hydro_text, "balance": self.row_balance_text, "tube": self.row_tube_text,
                                         "file_list": self.file_list}
                            with open(path_1, "w") as write_file:
                                json.dump(json_data, write_file, indent=4)
                            MyError.save_success(self)
                        else:
                            MyError.save_fail(self)
                else:
                    with open(path, "w") as write_file:
                        json.dump(json_data, write_file, indent=4)
                    if os.path.isfile(path):
                        self.comboBox_load_from.addItem(file_name_for_list)
                        self.file_list.append(file_name_for_list)
                        json_data = {"main": self.main_data, "heat": self.row_heat_text, "pr": self.row_pr_text,
                                     "hydro": self.row_hydro_text, "balance": self.row_balance_text, "tube": self.row_tube_text,
                                     "file_list": self.file_list}
                        with open(path_1, "w") as write_file:
                            json.dump(json_data, write_file, indent=4)
                        MyError.save_success(self)
                    else:
                        MyError.save_fail(self)
            else:
                return

    def json_load_data(self):
        self.pushButton_load_data.setDisabled(True)
        path_1 = os.path.join("data", "SVO_data.json")
        if self.comboBox_load_from.currentIndex() == -1:
            with open(path_1, "r") as read_file:
                self.json_data = json.load(read_file)
        else:
            file_name = self.comboBox_load_from.currentText() + '.json'
            path = os.path.join("data", file_name)
            with open(path, "r") as read_file:
                self.json_data = json.load(read_file)

    def load_main_data(self):
        try:
            self.main_data = self.json_data["main"]
            self.lineEdit_5_obj_name.setText(self.main_data[0])
            self.lineEdit_6_t_n.setText(self.main_data[1])
            self.lineEdit_11_pol_1.setText(self.main_data[2])
            self.lineEdit_13_pol_2.setText(self.main_data[3])
            self.lineEdit_7_pol_3.setText(self.main_data[4])
            self.lineEdit_14_pol_4.setText(self.main_data[5])
            self.lineEdit_9_ns_16.setText(self.main_data[6])
            self.lineEdit_9_ns_20.setText(self.main_data[7])
            self.lineEdit_15_kr.setText(self.main_data[8])
            self.lineEdit_12_wind.setText(self.main_data[9])
            self.lineEdit_16_door.setText(self.main_data[10])
            self.lineEdit_gate.setText(self.main_data[11])
            self.checkBox_3_n_coef.setChecked(self.main_data[12])
            self.lineEdit_p_n.setText(self.main_data[13])
            self.comboBox_t1.setCurrentIndex(self.main_data[14])
            self.comboBox_2_t2.setCurrentIndex(self.main_data[15])
            self.comboBox_4_p_n.setCurrentIndex(self.main_data[16])
            self.comboBox_3_tube.setCurrentIndex(self.main_data[17])
            self.lineEdit_17_sum_h_loss.setText(self.main_data[18])
            self.lineEdit_18_sum_heat.setText(self.main_data[19])
            self.lineEdit_20_sum_rate.setText(self.main_data[20])
            self.lineEdit_19_name_gck.setText(self.main_data[21])
            self.lineEdit_21_sum_p_loss.setText(self.main_data[22])
            self.lineEdit_22_sum_n_heat.setText(self.main_data[23])
            self.lineEdit_23_dl_heat.setText(self.main_data[24])
            self.lineEdit_24_n_rate.setText(self.main_data[25])
            self.lineEdit_25_dl_rate.setText(self.main_data[26])
            self.lineEdit_another_ok.setText(self.main_data[27])
            self.ns_room_long = self.main_data[28]
            self.lineEdit_s_n_sum.setText(self.main_data[29])
            self.lineEdit_s_iz_sum.setText(self.main_data[30])
            self.lineEdit_v_sum.setText(self.main_data[31])
            self.lineEdit_m_sum.setText(self.main_data[32])
        except:
            pass

    def load_heat_data(self):
        self.row_heat_text = self.json_data["heat"]
        for i in range(0, len(self.row_heat_text), 6):
            try:
                rowPosition = int(i / 6)
                if self.row_heat_text[i] == '':
                    self.add_room()
                    self.tableWidget.item(rowPosition, 1).setText(self.row_heat_text[i + 1])
                    self.tableWidget.item(rowPosition, 2).setText(self.row_heat_text[i + 2])
                    self.tableWidget.item(rowPosition, 4).setText(self.row_heat_text[i + 4])
                    self.tableWidget.item(rowPosition, 5).setText(self.row_heat_text[i + 5])
                elif self.row_heat_text[i] == 5:
                    self.add_row()
                    self.tableWidget.item(rowPosition, 2).setText(self.row_heat_text[i + 2])
                    self.tableWidget.item(rowPosition, 3).setText(self.row_heat_text[i + 3])
                    self.tableWidget.item(rowPosition, 4).setText(self.row_heat_text[i + 4])
                    self.tableWidget.item(rowPosition, 5).setText(self.row_heat_text[i + 5])
                    self.tableWidget.setCurrentCell(rowPosition, 0)
                    self.tableWidget.cellWidget(rowPosition, 0).setCurrentIndex(self.row_heat_text[i])
                    self.tableWidget.setCurrentCell(rowPosition, 1)
                    self.tableWidget.cellWidget(rowPosition, 1).setCurrentIndex(self.row_heat_text[i + 1])
                elif self.row_heat_text[i] == 6 or self.row_heat_text[i] == 7:
                    self.add_row()
                    self.tableWidget.item(rowPosition, 2).setText(self.row_heat_text[i + 2])
                    self.tableWidget.item(rowPosition, 3).setText(self.row_heat_text[i + 3])
                    self.tableWidget.item(rowPosition, 4).setText(self.row_heat_text[i + 4])
                    self.tableWidget.item(rowPosition, 5).setText(self.row_heat_text[i + 5])
                    self.tableWidget.setCurrentCell(rowPosition, 0)
                    self.tableWidget.cellWidget(rowPosition, 0).setCurrentIndex(self.row_heat_text[i])
                    self.tableWidget.setCurrentCell(rowPosition, 1)
                    self.tableWidget.cellWidget(rowPosition, 1).setCurrentIndex(self.row_heat_text[i + 1])
                    self.tableWidget.item(rowPosition + 1, 2).setText(self.row_heat_text[i + 8])
                    self.tableWidget.item(rowPosition + 1, 3).setText(self.row_heat_text[i + 9])
                    self.tableWidget.item(rowPosition + 1, 4).setText(self.row_heat_text[i + 10])
                    self.tableWidget.item(rowPosition + 1, 5).setText(self.row_heat_text[i + 11])
                elif self.row_heat_text[i] == 8:
                    self.add_row()
                    self.tableWidget.item(rowPosition, 2).setText(self.row_heat_text[i + 2])
                    self.tableWidget.item(rowPosition, 3).setText(self.row_heat_text[i + 3])
                    self.tableWidget.item(rowPosition, 4).setText(self.row_heat_text[i + 4])
                    self.tableWidget.item(rowPosition, 5).setText(self.row_heat_text[i + 5])
                    self.tableWidget.setCurrentCell(rowPosition, 0)
                    self.tableWidget.cellWidget(rowPosition, 0).setCurrentIndex(self.row_heat_text[i])
                    self.tableWidget.setCurrentCell(rowPosition, 1)
                    self.tableWidget.cellWidget(rowPosition, 1).setCurrentIndex(self.row_heat_text[i + 1])
                    self.tableWidget.item(rowPosition + 1, 2).setText(self.row_heat_text[i + 8])
                    self.tableWidget.item(rowPosition + 1, 3).setText(self.row_heat_text[i + 9])
                    self.tableWidget.item(rowPosition + 1, 4).setText(self.row_heat_text[i + 10])
                    self.tableWidget.item(rowPosition + 1, 5).setText(self.row_heat_text[i + 11])
                    self.tableWidget.item(rowPosition + 2, 2).setText(self.row_heat_text[i + 14])
                    self.tableWidget.item(rowPosition + 2, 3).setText(self.row_heat_text[i + 15])
                    self.tableWidget.item(rowPosition + 2, 4).setText(self.row_heat_text[i + 16])
                    self.tableWidget.item(rowPosition + 2, 5).setText(self.row_heat_text[i + 17])
                elif self.row_heat_text[i] == '' or self.row_heat_text[i] == '':
                    continue
                else:
                    self.add_row()
                    self.tableWidget.setCurrentCell(rowPosition, 0)
                    self.tableWidget.cellWidget(rowPosition, 0).setCurrentIndex(self.row_heat_text[i])
                    self.tableWidget.item(rowPosition, 2).setText(self.row_heat_text[i + 2])
                    self.tableWidget.item(rowPosition, 3).setText(self.row_heat_text[i + 3])
                    self.tableWidget.item(rowPosition, 4).setText(self.row_heat_text[i + 4])
                    self.tableWidget.item(rowPosition, 5).setText(self.row_heat_text[i + 5])
            except:
                pass

    def load_pr_data(self):
        try:
            self.row_pr_text = self.json_data["pr"]
            self.get_data_from_h_loss()
            if self.tableWidget_2.rowCount() == 0:
                self.pushButton_add_row.setDisabled(False)
                self.pushButton_3_del_row.setDisabled(False)
                self.pushButton_2_room.setDisabled(False)
                self.pushButton_4_calc_h_loss.setDisabled(False)
            self.comboBox_main_pr_type.setCurrentIndex(self.row_pr_text[-1])
            if self.comboBox_main_pr_type.currentIndex() == -1:
                self.delete_pr_table()
            for i in range(0, self.tableWidget_2.rowCount(), 5):
                self.rowPosition = i
                self.tableWidget_2.cellWidget(i, 4).setCurrentIndex(self.row_pr_text[i * 5 + 4])
                self.type_pr_choose()
                for m in range(5):
                    current_item = self.tableWidget_2.item(i + 2, m)
                    self.tableWidget_2.setCurrentCell(i + 2, m)
                    if isinstance(current_item, QtWidgets.QTableWidgetItem):
                        current_item.setText(self.row_pr_text[i * 5 + 10 + m])
                    else:
                        current_item = self.tableWidget_2.cellWidget(i + 2, m)
                        current_item.setCurrentIndex(self.row_pr_text[i * 5 + 10 + m])
                for m in range(5):
                    current_text = self.tableWidget_2.item(i + 4, m)
                    current_text.setText(self.row_pr_text[i * 5 + 20 + m])
        except:
            pass

    def load_hydro_data(self):
        self.row_hydro_text = self.json_data["hydro"]
        for i in range(0, len(self.row_hydro_text), self.tableWidget_3.rowCount()):
            colPosition = int(i / self.tableWidget_3.rowCount())
            self.add_section()
            for m in range(self.tableWidget_3.rowCount()):
                self.tableWidget_3.item(m, colPosition).setText(self.row_hydro_text[i + m])

    def load_balance_data(self):
        try:
            self.row_balance_text = self.json_data["balance"]
            self.vck = self.row_balance_text[-1:]
            self.vck = self.vck[0]
            self.row_balance_text.pop()
            self.get_data_from_hydro()
            for i in range(0, len(self.row_balance_text), self.tableWidget_4.rowCount()):
                colPosition = int(i / self.tableWidget_4.rowCount())
                for m in range(0, self.tableWidget_4.rowCount()):
                    self.tableWidget_4.item(m, colPosition).setText(self.row_balance_text[i + m])
            self.pushButton_calc_balance.setDisabled(True)
        except:
            pass

    def load_tube_data(self):
        self.row_tube_text = self.json_data["tube"]
        for i in range(0, len(self.row_tube_text), self.tableWidget_5.rowCount()):
            colPosition = int(i / self.tableWidget_5.rowCount())
            self.tube_add_column()
            for m in range(self.tableWidget_5.rowCount()):
                current_item = self.tableWidget_5.item(m, colPosition)
                self.tableWidget_5.setCurrentCell(m, colPosition)
                if isinstance(current_item, QtWidgets.QTableWidgetItem):
                    current_item.setText(self.row_tube_text[i + m])
                else:
                    current_item = self.tableWidget_5.cellWidget(m, colPosition)
                    current_item.setCurrentIndex(self.row_tube_text[i + m])

    def delete_files(self):
        MyError.error = QMessageBox()
        MyError.error.setWindowTitle(" ")
        MyError.error.setIcon(QMessageBox.Question)
        MyError.error.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        yesAll = MyError.error.button(QMessageBox.No)
        yesAll.setText(', ')
        noThis = MyError.error.button(QMessageBox.Yes)
        noThis.setText(' ')
        MyError.error.setText('     ?')
        path_icon = os.path.join("data", "icon.png")
        path_1 = os.path.join("data", "SVO_data.json")
        icon = QtGui.QIcon(path_icon)
        MyError.error.setWindowIcon(icon)
        del_confirm = MyError.error.exec_()
        if del_confirm == QMessageBox.No:
            MyError.error.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
            MyError.error.setText(
                '      .    '
                '?')
            len_file_list = self.comboBox_load_from.count()
            a = len_file_list
            del_confirm = MyError.error.exec_()
            if del_confirm == QMessageBox.Ok:
                for i in range(0, len_file_list):
                    a = a - 1
                    file_name = self.comboBox_load_from.itemText(a) + '.json'
                    file_name_from_list = self.comboBox_load_from.itemText(a)
                    path = os.path.join("data", file_name)
                    if os.path.isfile(path):
                        self.comboBox_load_from.removeItem(a)
                        self.file_list.remove(file_name_from_list)
                        json_data = {"main": '', "heat": '', "pr": '', "hydro": '', "balance": '', 'tube': '', "file_list": ''}
                        with open(path_1, "w") as write_file:
                            json.dump(json_data, write_file, indent=4)
                        os.remove(path)
            else:
                return
        elif del_confirm == QMessageBox.Yes:
            if self.comboBox_load_from.currentIndex() == -1:
                MyError.error.setStandardButtons(QMessageBox.Ok)
                MyError.error.setIcon(QMessageBox.Information)
                MyError.error.setText('    .')
                MyError.error.exec_()
                return
            else:
                MyError.error.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
                MyError.error.setText('  \{0}\?'.format(self.comboBox_load_from.currentText()))
                del_confirm = MyError.error.exec_()
                if del_confirm == QMessageBox.Ok:
                    a = self.comboBox_load_from.currentIndex()
                    file_name = self.comboBox_load_from.itemText(a) + '.json'
                    file_name_from_list = self.comboBox_load_from.itemText(a)
                    path = os.path.join("data", file_name)
                    if os.path.isfile(path):
                        self.comboBox_load_from.removeItem(a)
                        self.file_list.remove(file_name_from_list)
                        json_data = {"main": '', "heat": '', "pr": '', "hydro": '', "balance": '', 'tube': '',
                                     "file_list": self.file_list}
                        with open(path_1, "w") as write_file:
                            json.dump(json_data, write_file, indent=4)
                        os.remove(path)
                else:
                    return
        else:
            return

    def heat_data_to_excel(self):
        self.heat_text = []
        for i in range(self.tableWidget.columnCount()):
            for m in range(self.tableWidget.rowCount()):
                current_item = self.tableWidget.item(m, i)
                if isinstance(current_item, QtWidgets.QTableWidgetItem):
                    current_text = self.tableWidget.item(m, i).text()
                else:
                    current_text = self.tableWidget.cellWidget(m, i).currentText()
                self.heat_text.append(current_text)

    def pr_data_to_excel(self):
        self.pr_text = []
        for i in range(self.tableWidget_2.columnCount()):
            for m in range(self.tableWidget_2.rowCount()):
                current_item = self.tableWidget_2.item(m, i)
                if isinstance(current_item, QtWidgets.QTableWidgetItem):
                    current_text = self.tableWidget_2.item(m, i).text()
                else:
                    current_text = self.tableWidget_2.cellWidget(m, i).currentText()
                self.pr_text.append(current_text)

    def hydro_data_to_excel(self):
        self.hydro_param = []
        for i in range(self.tableWidget_3.rowCount()):
            self.hydro_param.append(self.tableWidget_3.verticalHeaderItem(i).text())
        self.hydro_excel_data = {'': self.hydro_param}
        self.hydro_text = []
        for i in range(self.tableWidget_3.columnCount()):
            for m in range(self.tableWidget_3.rowCount()):
                current_text = self.tableWidget_3.item(m, i).text()
                self.hydro_text.append(current_text)
            self.hydro_excel_data.update({i + 1: self.hydro_text})
            self.hydro_text = []

    def balance_data_to_excel(self):
        self.balance_text = []
        for i in range(self.tableWidget_4.columnCount()):
            for m in range(self.tableWidget_4.rowCount()):
                current_text = self.tableWidget_4.item(m, i).text()
                self.balance_text.append(current_text)

    def tube_data_to_excel(self):
        self.tube_param = []
        for i in range(self.tableWidget_5.rowCount()):
            self.tube_param.append(self.tableWidget_5.verticalHeaderItem(i).text())
        self.tube_excel_data = {'': self.tube_param}
        self.tube_text = []
        for i in range(self.tableWidget_5.columnCount()):
            for m in range(self.tableWidget_5.rowCount()):
                current_item = self.tableWidget_5.item(m, i)
                if isinstance(current_item, QtWidgets.QTableWidgetItem):
                    current_text = self.tableWidget_5.item(m, i).text()
                else:
                    current_text = self.tableWidget_5.cellWidget(m, i).currentText()
                self.tube_text.append(current_text)
            self.tube_excel_data.update({i + 1: self.tube_text})
            self.tube_text = []

    def import_data_to_excel(self):
        main_data = pd.DataFrame({self.lineEdit_5_obj_name.text(): [' ', ' -', ' I ',
                                                                    ' II ', ' III ', ' IV ',
                                                                    '  <16C', ' ',
                                                                    '', '', '',
                                                                    self.lineEdit_another_ok.text(),
                                                                    '- ',
                                                                    '- ', '  ',
                                                                    ' '],
                                  '1': ['', 't, C', 'k_pol_1, /2*C', 'k_pol_2, /2*C', 'k_pol_3, /2*C',
                                        'k_pol_4, /2*C', 'k_ns_<16, /2*C', 'k_ns, /2*C', 'k_kr, /2*C',
                                        'k_o, /2*C', 'k_d, /2*C', 'k_1, /2*C', 't1, C', 't2, C',
                                        'p1, {0}'.format(self.comboBox_4_p_n.currentText()), ''],
                                  '2': ['', self.lineEdit_6_t_n.text(), self.lineEdit_11_pol_1.text(),
                                        self.lineEdit_13_pol_2.text(), self.lineEdit_7_pol_3.text(),
                                        self.lineEdit_14_pol_4.text(), self.lineEdit_9_ns_16.text(),
                                        self.lineEdit_9_ns_20.text(), self.lineEdit_15_kr.text(),
                                        self.lineEdit_12_wind.text(), self.lineEdit_16_door.text(),
                                        self.lineEdit_gate.text(), self.comboBox_t1.currentText(),
                                        self.comboBox_2_t2.currentText(), self.lineEdit_p_n.text(),
                                        self.comboBox_3_tube.currentText()],
                                  '': ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
                                  '3': [' ', '  ', '  ',
                                        '  ', ' . ', '   ', '',
                                        '', '', '- ', '.  ', '  ',
                                        ' . ', ' . ', '', ''],
                                  '4': ['', 'Q, ', 'Q, ', 'G, /', '', 'p, ', '', '', '', '',
                                        'Q, ', 'G, /', 'Q, %', 'G, %', '', ''],
                                  '5': ['', self.lineEdit_17_sum_h_loss.text(), self.lineEdit_18_sum_heat.text(),
                                        self.lineEdit_20_sum_rate.text(), self.lineEdit_19_name_gck.text(),
                                        self.lineEdit_21_sum_p_loss.text(), '', '', '', '',
                                        self.lineEdit_22_sum_n_heat.text(), self.lineEdit_24_n_rate.text(),
                                        self.lineEdit_23_dl_heat.text(), self.lineEdit_25_dl_rate.text(), '', '']})

        heat_data = pd.DataFrame(
            {self.tableWidget.horizontalHeaderItem(0).text():
                 self.heat_text[:self.tableWidget.rowCount()],
             self.tableWidget.horizontalHeaderItem(1).text():
                 self.heat_text[self.tableWidget.rowCount():2 * self.tableWidget.rowCount()],
             self.tableWidget.horizontalHeaderItem(2).text():
                 self.heat_text[2 * self.tableWidget.rowCount():3 * self.tableWidget.rowCount()],
             self.tableWidget.horizontalHeaderItem(3).text():
                 self.heat_text[3 * self.tableWidget.rowCount():4 * self.tableWidget.rowCount()],
             self.tableWidget.horizontalHeaderItem(4).text():
                 self.heat_text[4 * self.tableWidget.rowCount():5 * self.tableWidget.rowCount()],
             self.tableWidget.horizontalHeaderItem(5).text():
                 self.heat_text[5 * self.tableWidget.rowCount():6 * self.tableWidget.rowCount()]})

        pr_data = pd.DataFrame(
            {self.tableWidget_2.horizontalHeaderItem(0).text():
                 self.pr_text[:self.tableWidget_2.rowCount()],
             self.tableWidget_2.horizontalHeaderItem(1).text():
                 self.pr_text[self.tableWidget_2.rowCount():2 * self.tableWidget_2.rowCount()],
             self.tableWidget_2.horizontalHeaderItem(2).text():
                 self.pr_text[2 * self.tableWidget_2.rowCount():3 * self.tableWidget_2.rowCount()],
             self.tableWidget_2.horizontalHeaderItem(3).text():
                 self.pr_text[3 * self.tableWidget_2.rowCount():4 * self.tableWidget_2.rowCount()],
             self.tableWidget_2.horizontalHeaderItem(4).text():
                 self.pr_text[4 * self.tableWidget_2.rowCount():5 * self.tableWidget_2.rowCount()]})

        hydro_data = pd.DataFrame(self.hydro_excel_data)

        balance_data = pd.DataFrame(
            {self.tableWidget_4.horizontalHeaderItem(0).text():
                 self.balance_text[:self.tableWidget_4.rowCount()],
             self.tableWidget_4.horizontalHeaderItem(1).text():
                 self.balance_text[self.tableWidget_4.rowCount():2 * self.tableWidget_4.rowCount()],
             self.tableWidget_4.horizontalHeaderItem(2).text():
                 self.balance_text[2 * self.tableWidget_4.rowCount():3 * self.tableWidget_4.rowCount()],
             self.tableWidget_4.horizontalHeaderItem(3).text():
                 self.balance_text[3 * self.tableWidget_4.rowCount():4 * self.tableWidget_4.rowCount()]})

        tube_data = pd.DataFrame(self.tube_excel_data)

        calc_so = {' ': main_data, '': heat_data, '': pr_data, '': hydro_data,
                   '': balance_data, '': tube_data}

        MyError.error = QMessageBox()
        MyError.error.setWindowTitle("   Excel")
        MyError.error.setIcon(QMessageBox.Question)
        MyError.error.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        path_icon = os.path.join("data", "icon.png")
        icon = QtGui.QIcon(path_icon)
        MyError.error.setWindowIcon(icon)
        if self.lineEdit_save_to_excel.text() == '':
            file_name = ' . ' + self.lineEdit_5_obj_name.text() + '.xlsx'
        else:
            file_name = ' . ' + self.lineEdit_5_obj_name.text() + '_' + self.lineEdit_save_to_excel.text() + '.xlsx'
        path = os.path.join("Excel", file_name)
        MyError.error.setText('    \{0}\?'.format(file_name))
        save_confirm = MyError.error.exec_()
        try:
            if save_confirm == QMessageBox.Ok:
                if os.path.isfile(path):
                    MyError.error.setText('     .   ?')
                    save_confirm = MyError.error.exec_()
                    if save_confirm == QMessageBox.Ok:
                        os.remove(path)
                        writer = pd.ExcelWriter(path, engine='openpyxl')
                        for sheet_name in calc_so.keys():
                            calc_so[sheet_name].to_excel(writer, sheet_name=sheet_name)
                        writer.close()
                        if os.path.isfile(path):
                            save_flag = True
                        else:
                            MyError.save_fail(self)
                else:
                    writer = pd.ExcelWriter(path, engine='openpyxl')
                    for sheet_name in calc_so.keys():
                        calc_so[sheet_name].to_excel(writer, sheet_name=sheet_name)
                    writer.close()
                    if os.path.isfile(path):
                        save_flag = True
                    else:
                        MyError.save_fail(self)
                if save_flag == True:
                    wb = pxl.load_workbook(path)
                    ws = wb.active
                    ws.column_dimensions['B'].width = 25
                    ws.column_dimensions['C'].width = 20
                    ws.column_dimensions['D'].width = 15
                    ws.column_dimensions['F'].width = 25
                    ws.column_dimensions['G'].width = 12
                    ws.column_dimensions['H'].width = 12
                    for r in range(1, 18):
                        for c in range(1, 9):
                            ws.cell(row=r, column=c).font = Font(name='Times New Roman', size=12)
                            ws.cell(row=r, column=c).border = Border(left=Side(border_style='thin'),
                                                                     right=Side(border_style='thin'),
                                                                     top=Side(border_style='thin'),
                                                                     bottom=Side(border_style='thin'))
                    for r in range(3, 18):
                        for c in range(3, 5):
                            ws.cell(row=r, column=c).alignment = Alignment(horizontal='center')
                    for r in range(3, 18):
                        ws.cell(row=r, column=8).alignment = Alignment(horizontal='center')
                    ws.merge_cells('B1:H1')
                    ws.merge_cells('B2:D2')
                    ws['B2'].alignment = Alignment(horizontal='center')
                    ws.merge_cells('F2:H2')
                    ws['F2'].alignment = Alignment(horizontal='center')
                    ws.merge_cells('F11:H11')
                    ws['F11'].alignment = Alignment(horizontal='center')
                    ws['B1'].font = Font(bold=True, name='Times New Roman', size=12)
                    wb.active = 1
                    ws1 = wb.active
                    ws1.column_dimensions['B'].width = 16
                    ws1.column_dimensions['C'].width = 14
                    ws1.column_dimensions['D'].width = 23
                    ws1.column_dimensions['E'].width = 16
                    ws1.column_dimensions['F'].width = 13
                    ws1.column_dimensions['G'].width = 15
                    ws1.row_dimensions[1].height = 30
                    for r in range(1, self.tableWidget.rowCount() + 2):
                        for c in range(1, 8):
                            ws1.cell(row=r, column=c).font = Font(name='Times New Roman', size=11)
                            ws1.cell(row=r, column=c).border = Border(left=Side(border_style='thin'),
                                                                      right=Side(border_style='thin'),
                                                                      top=Side(border_style='thin'),
                                                                      bottom=Side(border_style='thin'))
                            ws1.cell(row=r, column=c).alignment = Alignment(horizontal='center')
                    for c in range(1, 8):
                        ws1.cell(row=1, column=c).font = Font(bold=True, name='Times New Roman', size=11)
                        ws1.cell(row=1, column=c).alignment = Alignment(horizontal='center', vertical='center',
                                                                        wrap_text=True)
                    for r in range(1, self.tableWidget.rowCount() + 2):
                        if ws1.cell(row=r, column=2).value == '':
                            for c in range(1, 8):
                                ws1.cell(row=r, column=c).fill = PatternFill('solid', fgColor='7FFFD4')
                    wb.active = 2
                    ws2 = wb.active
                    ws2.column_dimensions['B'].width = 18
                    ws2.column_dimensions['C'].width = 10
                    ws2.column_dimensions['D'].width = 25
                    ws2.column_dimensions['E'].width = 16
                    ws2.column_dimensions['F'].width = 20
                    for r in range(1, self.tableWidget_2.rowCount() + 2):
                        for c in range(1, 7):
                            ws2.cell(row=r, column=c).font = Font(name='Times New Roman', size=11)
                            ws2.cell(row=r, column=c).border = Border(left=Side(border_style='thin'),
                                                                      right=Side(border_style='thin'),
                                                                      top=Side(border_style='thin'),
                                                                      bottom=Side(border_style='thin'))
                            ws2.cell(row=r, column=c).alignment = Alignment(horizontal='center')
                    for r in range(1, self.tableWidget_2.rowCount() + 2):
                        if ws2.cell(row=r, column=2).value == '':
                            for c in range(1, 7):
                                ws2.cell(row=r, column=c).fill = PatternFill('solid', fgColor='7FFFD4')
                    wb.active = 3
                    ws3 = wb.active
                    ws3.column_dimensions['B'].width = 18
                    for r in range(1, 20):
                        for c in range(1, self.tableWidget_3.columnCount() + 3):
                            ws3.cell(row=r, column=c).font = Font(name='Times New Roman', size=11)
                            ws3.cell(row=r, column=c).border = Border(left=Side(border_style='thin'),
                                                                      right=Side(border_style='thin'),
                                                                      top=Side(border_style='thin'),
                                                                      bottom=Side(border_style='thin'))
                            ws3.cell(row=r, column=c).alignment = Alignment(horizontal='center')
                    for c in range(3, self.tableWidget_3.columnCount() + 3):
                        ws3.cell(row=2, column=c).fill = PatternFill('solid', fgColor='00BFFF')
                        ws3.cell(row=12, column=c).fill = PatternFill('solid', fgColor='00BFFF')
                        if float(ws3.cell(row=13, column=c).value) > 0.6:
                            ws3.cell(row=13, column=c).fill = PatternFill('solid', fgColor='F08080')
                            ws3.cell(row=4, column=c).fill = PatternFill('solid', fgColor='F08080')
                    try:
                        ws3.cell(row=19, column=self.gck_ind + 3).fill = PatternFill('solid', fgColor='FFD700')
                        for i in range(len(self.vck)):
                            ws3.cell(row=19, column=self.vck[i] + 3).fill = PatternFill('solid', fgColor='00FF7F')
                    except:
                        pass
                    wb.active = 4
                    ws4 = wb.active
                    ws4.column_dimensions['D'].width = 12
                    ws4.column_dimensions['E'].width = 15
                    for r in range(1, self.tableWidget_4.rowCount() + 2):
                        for c in range(1, 6):
                            ws4.cell(row=r, column=c).font = Font(name='Times New Roman', size=11)
                            ws4.cell(row=r, column=c).border = Border(left=Side(border_style='thin'),
                                                                      right=Side(border_style='thin'),
                                                                      top=Side(border_style='thin'),
                                                                      bottom=Side(border_style='thin'))
                            ws4.cell(row=r, column=c).alignment = Alignment(horizontal='center')
                    for c in range(1, 6):
                        ws4.cell(row=1, column=c).font = Font(bold=True, name='Times New Roman', size=11)
                        ws4.cell(row=1, column=c).alignment = Alignment(horizontal='center', vertical='center',
                                                                        wrap_text=True)
                    for r in range(2, self.tableWidget_4.rowCount() + 2):
                        if float(ws4.cell(row=r, column=4).value) <= 3:
                            ws4.cell(row=r, column=4).fill = PatternFill('solid', fgColor='F08080')
                        if float(ws4.cell(row=r, column=5).value) < 1:
                            ws4.cell(row=r, column=5).fill = PatternFill('solid', fgColor='F08080')
                    wb.active = 5
                    ws5 = wb.active
                    ws5.column_dimensions['B'].width = 17
                    for r in range(1, 19):
                        for c in range(1, self.tableWidget_5.columnCount() + 3):
                            ws5.cell(row=r, column=c).font = Font(name='Times New Roman', size=11)
                            ws5.cell(row=r, column=c).border = Border(left=Side(border_style='thin'),
                                                                      right=Side(border_style='thin'),
                                                                      top=Side(border_style='thin'),
                                                                      bottom=Side(border_style='thin'))
                            ws5.cell(row=r, column=c).alignment = Alignment(horizontal='center')
                    for c in range(3, self.tableWidget_5.columnCount() + 3):
                        ws5.cell(row=2, column=c).fill = PatternFill('solid', fgColor='7FFFD4')
                        ws5.cell(row=3, column=c).fill = PatternFill('solid', fgColor='7FFFD4')
                        ws5.cell(row=9, column=c).fill = PatternFill('solid', fgColor='00BFFF')
                        ws5.cell(row=12, column=c).fill = PatternFill('solid', fgColor='00BFFF')
                        ws5.cell(row=13, column=c).fill = PatternFill('solid', fgColor='00BFFF')
                        if ws5.cell(row=2, column=c).value == '  ' and float(ws5.cell(row=9, column=c).value) > 0.6:
                            ws5.cell(row=9, column=c).fill = PatternFill('solid', fgColor='F08080')
                            ws5.cell(row=10, column=c).fill = PatternFill('solid', fgColor='F08080')
                            ws5.cell(row=11, column=c).fill = PatternFill('solid', fgColor='F08080')
                        elif float(ws5.cell(row=9, column=c).value) > 1.2:
                            ws5.cell(row=9, column=c).fill = PatternFill('solid', fgColor='F08080')
                            ws5.cell(row=10, column=c).fill = PatternFill('solid', fgColor='F08080')
                            ws5.cell(row=11, column=c).fill = PatternFill('solid', fgColor='F08080')
                        for i in range(self.tableWidget_5.columnCount()):
                            if self.comboBox_q_r_choose.currentIndex() == 0 and self.tableWidget_5.cellWidget(0, i).currentIndex() == 0:
                                ws5.cell(row=4, column=i+3).fill = PatternFill('solid', fgColor='00BFFF')
                            elif self.comboBox_q_r_choose.currentIndex() == 1 and self.tableWidget_5.cellWidget(0, i).currentIndex() == 0:
                                ws5.cell(row=5, column=i+3).fill = PatternFill('solid', fgColor='00BFFF')
                            elif self.tableWidget_5.cellWidget(0, i).currentIndex() != 0:
                                ws5.cell(row=7, column=i+3).fill = PatternFill('solid', fgColor='00BFFF')
                            letter = get_column_letter(i+3)
                            ws5.column_dimensions[letter].width = 14
                    for r in range(20, 24):
                        for c in range(1, 4):
                            ws5.cell(row=r, column=c).font = Font(name='Times New Roman', size=11)
                            ws5.cell(row=r, column=c).border = Border(left=Side(border_style='thin'),
                                                                      right=Side(border_style='thin'),
                                                                      top=Side(border_style='thin'),
                                                                      bottom=Side(border_style='thin'))
                            ws5.cell(row=r, column=c).alignment = Alignment(horizontal='center')
                    for r in range(20, 24):
                        ws5.cell(row=r, column=1).value = r - 3
                    ws5.cell(row=20, column=2).value = "S, 2"
                    ws5.cell(row=21, column=2).value = "S, 2"
                    ws5.cell(row=22, column=2).value = "v, 3"
                    ws5.cell(row=23, column=2).value = "M, "
                    ws5.cell(row=20, column=3).value = self.lineEdit_s_n_sum.text()
                    ws5.cell(row=21, column=3).value = self.lineEdit_s_iz_sum.text()
                    ws5.cell(row=22, column=3).value = self.lineEdit_v_sum.text()
                    ws5.cell(row=23, column=3).value = self.lineEdit_m_sum.text()
                    wb.active = 0
                    wb.save(path)
                    MyError.save_success(self)
                return
        except PermissionError:
            MyError.save_fail(self)

    #   
    def manual_ex(self):
        self.Manual.show()

    def change_style(self):
        self.tableWidget.setAlternatingRowColors(False)
        self.tableWidget_2.setAlternatingRowColors(False)
        self.tableWidget_3.setAlternatingRowColors(False)
        self.tableWidget_4.setAlternatingRowColors(False)
        self.tableWidget_5.setAlternatingRowColors(False)
        if self.comboBox_wind_style.currentIndex() == 0:
            self.setStyleSheet("background-color: rgb(200, 200, 200);\n"
                               "font: 12pt \"Times New Roman\";")
            self.Manual.setStyleSheet("background-color: rgb(200, 200, 200);\n"
                                      "font: 12pt \"Times New Roman\";")
        if self.comboBox_wind_style.currentIndex() == 1:
            self.setStyleSheet("background-color: rgb(215, 215, 215);\n"
                               "font: 12pt \"Times New Roman\";")
            self.Manual.setStyleSheet("background-color: rgb(215, 215, 215);\n"
                                      "font: 12pt \"Times New Roman\";")
        if self.comboBox_wind_style.currentIndex() == 2:
            self.setStyleSheet("background-color: rgb(230, 230, 230);\n"
                               "font: 12pt \"Times New Roman\";")
            self.Manual.setStyleSheet("background-color: rgb(230, 230, 230);\n"
                                      "font: 12pt \"Times New Roman\";")


if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    w = MainWindow()
    w.show()
    sys.exit(app.exec_())
